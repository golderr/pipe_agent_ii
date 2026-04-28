from __future__ import annotations

import logging
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook
from typer.testing import CliRunner

from tcg_pipeline.cli import app
from tcg_pipeline.db.models import (
    DismissReason,
    IdentifierType,
    PipelineStatus,
    RelationshipType,
)
from tcg_pipeline.ingesters.pipedream import PipedreamIngester
from tcg_pipeline.permit_numbers import extract_ladbs_pcis_permit_numbers

runner = CliRunner()


def test_extract_ladbs_pcis_permit_numbers_ignores_malformed_urls_and_dedupes_matches() -> None:
    permit_numbers = extract_ladbs_pcis_permit_numbers(
        [
            (
                "First https://www.ladbsservices2.lacity.org/OnlineServices/PermitReport/"
                "PcisPermitDetail?id1=18010&id2=10000&id3=03620#details "
                "second https://www.ladbsservices2.lacity.org/OnlineServices/PermitReport/"
                "PcisPermitDetail?id1=22010&id2=10000&id3=04890"
            ),
            (
                "https://www.ladbsservices2.lacity.org/OnlineServices/PermitReport/"
                "PcisPermitDetail?id1=18010&id2=10000&id3=03620"
            ),
            (
                "https://www.ladbsservices2.lacity.org/OnlineServices/PermitReport/"
                "PcisPermitDetail?id1=bad&id2=10000&id3=03620"
            ),
            (
                "https://www.ladbsservices2.lacity.org/OnlineServices/PermitReport/"
                "PcisPermitDetail?id1=23010&id2=10000"
            ),
            "https://example.com/not-ladbs",
        ]
    )

    assert permit_numbers == [
        "18010-10000-03620",
        "22010-10000-04890",
    ]


def test_ingest_workbook_builds_project_records_and_history(tmp_path: Path) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_sample.xlsx",
        [
            {
                "ProjectID": "23.00001",
                "Name": "Palladium Residences",
                "Developer": "CIM Group",
                "Address": "5939 W Sunset Blvd",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90028",
                "Region": "Hollywood/Los Feliz, CA",
                "Lat": 34.0972,
                "Long": -118.3201,
                "RentFS": "Rental",
                "MRUnits": 150,
                "AffUnits": 25,
                "TotUnits": 175,
                "Acres": 1.5,
                "RetailSF": 5000,
                "ProdType": "Apartment",
                "Elevation": 7,
                "Senior": "Non Age-Restricted",
                "PercS": 0.15,
                "Perc1B": 0.45,
                "Perc2B": 0.35,
                "PercOther": 0.05,
                "CurrStatus": "Pending",
                "CurrStatusDate": date(2026, 4, 1),
                "RefNum": "DIR-2024-1234-TOC",
                "APN": 5544027012,
                "Notes": "Survived appeal.",
                "Site1": "https://planning.lacity.gov/case",
                "Site2": "https://developer.example.com/project",
                "PersonalNotes": "Follow up next cycle.",
                "ChangeNotes": "Added updated unit count.",
                "PStat1": "Proposed",
                "PStatDate1": date(2025, 12, 15),
                "PStat2": "Conceptual",
                "PStatDate2": date(2025, 6, 1),
                "PrevName1": "Sunset Palladium",
                "RelP1": "23.00002",
                "DeliveryDate": date(2028, 1, 1),
                "Editor": "NG",
                "EditDate": date(2026, 4, 10),
            }
        ],
    )

    result = PipedreamIngester(market="los_angeles").ingest_workbook(workbook_path)

    assert result.imported_count == 1
    assert result.dismissed_count == 0
    assert len(result.staged_relationships) == 1

    record = result.project_records[0]
    project = record.project

    assert record.project_identifier_value == "23.00001"
    assert project.project_name == "Palladium Residences"
    assert project.developer == "CIM Group"
    assert project.canonical_address == "5939 WEST SUNSET BOULEVARD LOS ANGELES CA 90028"
    assert project.raw_addresses == ["5939 W Sunset Blvd"]
    assert project.pipeline_status == PipelineStatus.PENDING
    assert project.source_urls == [
        "https://planning.lacity.gov/case",
        "https://developer.example.com/project",
    ]
    assert project.previous_names == ["Sunset Palladium"]
    assert project.total_units == 175
    assert project.lat == 34.0972
    assert project.lng == -118.3201
    assert project.status_date == date(2026, 4, 1)
    assert project.last_editor == "NG"
    assert project.last_edit_date == date(2026, 4, 10)
    assert {
        note.note_type: note.body
        for note in record.project_notes
    } == {
        "researcher_notes": "Survived appeal.",
        "personal_notes": "Follow up next cycle.",
        "change_notes": "Added updated unit count.",
    }

    identifier_types = {identifier.identifier_type for identifier in record.identifiers}
    assert IdentifierType.TCG_PIPEDREAM_ID in identifier_types
    assert IdentifierType.CASE_NUMBER in identifier_types
    assert IdentifierType.APN in identifier_types

    statuses = [status_entry.status for status_entry in record.status_history]
    assert statuses == [
        PipelineStatus.CONCEPTUAL,
        PipelineStatus.PROPOSED,
        PipelineStatus.PENDING,
    ]
    assert record.source_record.source_record_id == "23.00001"
    assert record.source_record.source_url == "https://planning.lacity.gov/case"

    staged_relationship = result.staged_relationships[0]
    assert staged_relationship.project_identifier_value == "23.00001"
    assert staged_relationship.related_project_identifier_value == "23.00002"
    assert staged_relationship.relationship_type == RelationshipType.PHASE


def test_ingest_workbook_dismisses_delete_records_and_preserves_duplicate_link(
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_delete.xlsx",
        [
            {
                "ProjectID": 23.00003,
                "Address": "1718 N Las Palmas Ave",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90028",
                "CurrStatus": "Delete - Duplicate",
                "CorrP": 23.1,
            }
        ],
    )

    result = PipedreamIngester(market="los_angeles").ingest_workbook(workbook_path)

    assert result.imported_count == 0
    assert result.dismissed_count == 1
    assert len(result.staged_relationships) == 1

    dismissed = result.dismissed_records[0]
    assert dismissed.source_record_id == "23.00003"
    assert dismissed.reason == DismissReason.DUPLICATE
    assert dismissed.notes == "Delete - Duplicate; CorrP=23.10000"

    staged_relationship = result.staged_relationships[0]
    assert staged_relationship.project_identifier_value == "23.00003"
    assert staged_relationship.related_project_identifier_value == "23.10000"
    assert staged_relationship.relationship_type == RelationshipType.DUPLICATE


def test_ingest_workbook_can_filter_to_allowed_city(tmp_path: Path) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_city_filter.xlsx",
        [
            {
                "ProjectID": "23.00001",
                "Address": "5939 W Sunset Blvd",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Pending",
            },
            {
                "ProjectID": "88.00001",
                "Address": "9000 Sunset Blvd",
                "State": "CA",
                "County": "Los Angeles",
                "City": "West Hollywood",
                "CurrStatus": "Pending",
            },
        ],
    )

    result = PipedreamIngester(
        market="los_angeles",
        allowed_cities=["Los Angeles"],
    ).ingest_workbook(workbook_path)

    assert result.imported_count == 1
    assert result.skipped_project_ids == ["88.00001"]


def test_ingest_workbook_uses_name_as_address_when_address_is_entity_name(
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_address_fallback.xlsx",
        [
            {
                "ProjectID": "24.00085",
                "Name": "8000 W 3rd St",
                "Developer": "Jon Bosse",
                "Address": "8008 Third Street Investments, LLC",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": 90048,
                "CurrStatus": "Inactive",
            }
        ],
    )

    result = PipedreamIngester(market="los_angeles").ingest_workbook(workbook_path)

    assert result.imported_count == 1
    project = result.project_records[0].project
    assert project.canonical_address == "8000 WEST 3RD STREET LOS ANGELES CA 90048"
    assert project.state == "CA"
    assert project.raw_addresses == [
        "8000 W 3rd St",
        "8008 Third Street Investments, LLC",
    ]


def test_ingest_workbook_does_not_harvest_ladbs_permit_numbers_outside_los_angeles_market(
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_non_la_market.xlsx",
        [
            {
                "ProjectID": "24.00086",
                "Address": "1437 4th Street",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Santa Monica",
                "CurrStatus": "Pending",
                "Site1": (
                    "https://www.ladbsservices2.lacity.org/OnlineServices/PermitReport/"
                    "PcisPermitDetail?id1=18010&id2=10000&id3=03620"
                ),
            }
        ],
    )

    result = PipedreamIngester(market="santa_monica").ingest_workbook(workbook_path)

    assert result.imported_count == 1
    identifier_types = {
        identifier.identifier_type for identifier in result.project_records[0].identifiers
    }
    assert IdentifierType.PERMIT_NUMBER not in identifier_types


def test_ingest_workbook_extracts_pcis_permit_identifiers_from_source_urls(
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_permit_urls.xlsx",
        [
            {
                "ProjectID": "24.00223",
                "Address": "10608 W Pico Blvd",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Approved",
                "Site1": (
                    "https://www.ladbsservices2.lacity.org/OnlineServices/PermitReport/"
                    "PcisPermitDetail?id1=22010&id2=10000&id3=04890"
                ),
                "Site2": (
                    "https://www.ladbsservices2.lacity.org/OnlineServices/PermitReport/"
                    "PcisPermitDetail?id1=22010&id2=10000&id3=04890"
                ),
            }
        ],
    )

    result = PipedreamIngester(market="los_angeles").ingest_workbook(workbook_path)

    assert result.imported_count == 1
    identifiers = {
        (identifier.identifier_type, identifier.value)
        for identifier in result.project_records[0].identifiers
    }
    assert (IdentifierType.PERMIT_NUMBER, "22010-10000-04890") in identifiers


def test_preview_pipedream_command_reports_counts(tmp_path: Path) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_cli.xlsx",
        [
            {
                "ProjectID": "23.00001",
                "Address": "5939 W Sunset Blvd",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Pending",
            }
        ],
    )

    result = runner.invoke(
        app,
        ["preview-pipedream", str(workbook_path), "--market", "los_angeles"],
    )

    assert result.exit_code == 0
    assert "Imported projects: 1" in result.stdout
    assert "Dismissed records: 0" in result.stdout
    assert "Missing ProjectID rows: 0" in result.stdout


def test_ingest_workbook_records_diagnostics_for_invalid_values(
    tmp_path: Path,
    caplog: pytest.LogCaptureFixture,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_invalid.xlsx",
        [
            {
                "ProjectID": "23.00001",
                "Address": "5939 W Sunset Blvd",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "04/01/2026",
                "CurrStatusDate": "not a date",
                "RentFS": "Lease",
                "ProdType": "Tower",
                "Senior": "Retiree",
            },
            {
                "ProjectID": 23.1,
                "Address": "9000 Sunset Blvd",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Pending",
            },
        ],
    )

    caplog.set_level(logging.WARNING)
    result = PipedreamIngester(market="los_angeles").ingest_workbook(workbook_path)

    assert result.imported_count == 2
    assert result.issue_counts == {
        "invalid_date": 1,
        "invalid_enum": 3,
        "invalid_status": 1,
        "suspicious_identifier": 1,
    }
    assert result.project_records[0].project.pipeline_status == PipelineStatus.PROPOSED
    assert result.project_records[1].project_identifier_value == "23.10000"
    assert "Unrecognized pipeline status; defaulting to Proposed" in caplog.text
    assert "Project identifier normalized to '23.10000'" in caplog.text


def _build_pipedream_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "DataStorage"

    headers = _ordered_headers(rows)
    for index, header in enumerate(headers, start=1):
        column = (index * 2) - 1
        worksheet.cell(row=3, column=column, value=header)
        for row_index, row_values in enumerate(rows, start=4):
            if header in row_values:
                worksheet.cell(row=row_index, column=column, value=row_values[header])

    workbook.save(path)
    workbook.close()
    return path


def _ordered_headers(rows: list[dict[str, object]]) -> list[str]:
    preferred_order = [
        "ProjectID",
        "Name",
        "Developer",
        "Address",
        "State",
        "County",
        "City",
        "Zip",
        "Region",
        "Lat",
        "Long",
        "RentFS",
        "MRUnits",
        "AffUnits",
        "TotUnits",
        "Acres",
        "RetailSF",
        "ProdType",
        "Elevation",
        "Senior",
        "PercS",
        "Perc1B",
        "Perc2B",
        "PercOther",
        "CurrStatus",
        "CurrStatusDate",
        "RefNum",
        "APN",
        "Notes",
        "Site1",
        "Site2",
        "PersonalNotes",
        "ChangeNotes",
        "PStat1",
        "PStatDate1",
        "PStat2",
        "PStatDate2",
        "PrevName1",
        "RelP1",
        "CorrP",
        "DeliveryDate",
        "Editor",
        "EditDate",
    ]
    seen = set(preferred_order)
    dynamic_headers = []
    for row in rows:
        for header in row:
            if header not in seen:
                dynamic_headers.append(header)
                seen.add(header)
    return preferred_order + dynamic_headers
