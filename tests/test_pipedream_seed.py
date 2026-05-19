from __future__ import annotations

from contextlib import contextmanager
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session
from typer.testing import CliRunner

from tcg_pipeline.cli import app, resolve_all
from tcg_pipeline.db.models import (
    ChangeLog,
    Evidence,
    IdentifierType,
    Project,
    ProjectIdentifier,
    ProjectRelationship,
    ProjectSourceRecord,
    ResolutionLog,
    StatusHistory,
)
from tcg_pipeline.db.seed import (
    ingest_pipedream_workbooks,
    persist_pipedream_import_results,
)

runner = CliRunner()


def test_persist_import_results_resolves_cross_file_relationships(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    workbook_one = _build_pipedream_workbook(
        tmp_path / "pipedream_one.xlsx",
        [
            {
                "ProjectID": "991.00001",
                "Address": "9901 W Example Blvd",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Pending",
                "RelP1": "992.00001",
            }
        ],
    )
    workbook_two = _build_pipedream_workbook(
        tmp_path / "pipedream_two.xlsx",
        [
            {
                "ProjectID": "992.00001",
                "Address": "9902 N Example Ave",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Approved",
            }
        ],
    )

    import_results = ingest_pipedream_workbooks(
        [workbook_one, workbook_two],
        market="los_angeles",
    )
    persist_result = persist_pipedream_import_results(postgres_session, import_results)

    assert persist_result.inserted_projects == 2
    assert persist_result.created_relationships == 1
    assert persist_result.unresolved_relationship_count == 0

    test_ids = ["991.00001", "992.00001"]
    identifier_rows = postgres_session.execute(
        select(ProjectIdentifier.value).where(
            ProjectIdentifier.identifier_type == IdentifierType.TCG_PIPEDREAM_ID,
            ProjectIdentifier.value.in_(test_ids),
        )
    ).scalars()
    assert sorted(identifier_rows) == test_ids

    source_project_id = postgres_session.execute(
        select(ProjectIdentifier.project_id).where(
            ProjectIdentifier.identifier_type == IdentifierType.TCG_PIPEDREAM_ID,
            ProjectIdentifier.value == "991.00001",
        )
    ).scalar_one()
    relationship = postgres_session.execute(
        select(ProjectRelationship).where(
            ProjectRelationship.project_id == source_project_id,
        )
    ).scalar_one()
    assert relationship.relationship_type.value == "phase"


def test_persist_import_results_reports_unresolved_relationships(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_unresolved.xlsx",
        [
            {
                "ProjectID": "993.00001",
                "Address": "9903 S Example Dr",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Pending",
                "RelP1": "999.00001",
            }
        ],
    )

    import_results = ingest_pipedream_workbooks([workbook_path], market="los_angeles")
    persist_result = persist_pipedream_import_results(postgres_session, import_results)

    assert persist_result.inserted_projects == 1
    assert persist_result.created_relationships == 0
    assert persist_result.unresolved_relationship_count == 1
    assert persist_result.unresolved_relationships[0].missing_identifiers == ["999.00001"]


def test_seed_pipedream_command_persists_when_not_dry_run(
    monkeypatch: pytest.MonkeyPatch,
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_cli_seed.xlsx",
        [
            {
                "ProjectID": "994.00001",
                "Address": "9904 E Example Ln",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Pending",
            }
        ],
    )

    @contextmanager
    def fake_session_factory() -> Session:
        yield postgres_session

    monkeypatch.setattr("tcg_pipeline.cli.get_session_factory", lambda: fake_session_factory)

    result = runner.invoke(
        app,
        ["seed-pipedream", str(workbook_path), "--market", "los_angeles"],
    )

    assert result.exit_code == 0
    assert "Persisted projects: 1" in result.stdout

    persisted_identifier = postgres_session.execute(
        select(ProjectIdentifier.value).where(
            ProjectIdentifier.identifier_type == IdentifierType.TCG_PIPEDREAM_ID,
            ProjectIdentifier.value == "994.00001",
        )
    ).scalar_one()
    assert persisted_identifier == "994.00001"


def test_seed_pipedream_with_defer_resolve_skips_resolution_writes(
    monkeypatch: pytest.MonkeyPatch,
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_defer_resolve.xlsx",
        [
            {
                "ProjectID": "994.10001",
                "Address": "9914 E Deferred Ln",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Under Construction",
                "CurrStatusDate": "2026-04-01",
                "TotUnits": 120,
            }
        ],
    )

    @contextmanager
    def fake_session_factory() -> Session:
        yield postgres_session

    monkeypatch.setattr("tcg_pipeline.cli.get_session_factory", lambda: fake_session_factory)

    result = runner.invoke(
        app,
        ["seed-pipedream", str(workbook_path), "--market", "los_angeles", "--defer-resolve"],
    )

    assert result.exit_code == 0
    assert "Deferred inline resolve: True" in result.stdout
    project_id = postgres_session.execute(
        select(ProjectIdentifier.project_id).where(
            ProjectIdentifier.identifier_type == IdentifierType.TCG_PIPEDREAM_ID,
            ProjectIdentifier.value == "994.10001",
        )
    ).scalar_one()
    project = postgres_session.get(Project, project_id)

    assert project is not None
    assert project.last_evidence_date is None
    assert _count_rows(postgres_session, Evidence) == 1
    assert _count_rows(postgres_session, ProjectSourceRecord) == 1
    assert _count_rows(postgres_session, StatusHistory) == 1
    assert _count_rows(postgres_session, ResolutionLog) == 0
    assert _count_rows(postgres_session, ChangeLog) == 0


def test_seed_pipedream_without_flag_resolves_inline(
    monkeypatch: pytest.MonkeyPatch,
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_inline_resolve.xlsx",
        [
            {
                "ProjectID": "994.10002",
                "Address": "9915 E Inline Ln",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Under Construction",
                "CurrStatusDate": "2026-04-01",
                "TotUnits": 121,
            }
        ],
    )

    @contextmanager
    def fake_session_factory() -> Session:
        yield postgres_session

    monkeypatch.setattr("tcg_pipeline.cli.get_session_factory", lambda: fake_session_factory)

    result = runner.invoke(
        app,
        ["seed-pipedream", str(workbook_path), "--market", "los_angeles"],
    )

    assert result.exit_code == 0
    assert "Deferred inline resolve: False" in result.stdout
    project_id = postgres_session.execute(
        select(ProjectIdentifier.project_id).where(
            ProjectIdentifier.identifier_type == IdentifierType.TCG_PIPEDREAM_ID,
            ProjectIdentifier.value == "994.10002",
        )
    ).scalar_one()
    project = postgres_session.get(Project, project_id)

    assert project is not None
    assert project.last_evidence_date is not None
    assert _count_rows(postgres_session, ResolutionLog) > 0


def test_resolve_all_apply_completes_deferred_pipedream_seed_state(
    monkeypatch: pytest.MonkeyPatch,
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_deferred_resolve_all.xlsx",
        [
            {
                "ProjectID": "994.10003",
                "Address": "9916 E Resolve All Ln",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Under Construction",
                "CurrStatusDate": "2026-04-01",
                "TotUnits": 122,
            }
        ],
    )

    @contextmanager
    def fake_session_factory() -> Session:
        yield postgres_session

    monkeypatch.setattr("tcg_pipeline.cli.get_session_factory", lambda: fake_session_factory)

    seed_result = runner.invoke(
        app,
        ["seed-pipedream", str(workbook_path), "--market", "los_angeles", "--defer-resolve"],
    )
    assert seed_result.exit_code == 0

    project_id = postgres_session.execute(
        select(ProjectIdentifier.project_id).where(
            ProjectIdentifier.identifier_type == IdentifierType.TCG_PIPEDREAM_ID,
            ProjectIdentifier.value == "994.10003",
        )
    ).scalar_one()
    project = postgres_session.get(Project, project_id)
    assert project is not None
    assert project.last_evidence_date is None
    assert _count_rows(postgres_session, ResolutionLog) == 0
    assert _count_rows(postgres_session, StatusHistory) == 1

    resolve_all(market="los_angeles", apply=True, clear_log=True, limit=1, batch_size=1)
    postgres_session.expire_all()
    project = postgres_session.get(Project, project_id)

    assert project is not None
    assert project.last_evidence_date is not None
    assert project.confidence_reason is not None
    assert project.likelihood is not None
    assert _count_rows(postgres_session, ResolutionLog) > 0
    assert _count_rows(postgres_session, StatusHistory) == 1


def test_persist_import_results_is_idempotent_for_existing_project_ids(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_repeat.xlsx",
        [
            {
                "ProjectID": "995.00001",
                "Address": "9905 W Example Ct",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Pending",
            }
        ],
    )

    first_import_results = ingest_pipedream_workbooks([workbook_path], market="los_angeles")
    second_import_results = ingest_pipedream_workbooks([workbook_path], market="los_angeles")
    first_persist_result = persist_pipedream_import_results(
        postgres_session,
        first_import_results,
    )
    second_persist_result = persist_pipedream_import_results(
        postgres_session,
        second_import_results,
    )

    assert first_persist_result.inserted_projects == 1
    assert second_persist_result.inserted_projects == 0
    assert second_persist_result.skipped_existing_project_ids == ["995.00001"]

    identifier_row = postgres_session.execute(
        select(ProjectIdentifier.value).where(
            ProjectIdentifier.identifier_type == IdentifierType.TCG_PIPEDREAM_ID,
            ProjectIdentifier.value == "995.00001",
        )
    ).scalar_one()
    assert identifier_row == "995.00001"


def test_persist_import_results_stores_harvested_permit_number_identifiers(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_permit_identifier.xlsx",
        [
            {
                "ProjectID": "996.00001",
                "Address": "329 S Bonnie Brae St",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "CurrStatus": "Approved",
                "Site1": (
                    "https://www.ladbsservices2.lacity.org/OnlineServices/PermitReport/"
                    "PcisPermitDetail?id1=18010&id2=10000&id3=03620"
                ),
            }
        ],
    )

    import_results = ingest_pipedream_workbooks([workbook_path], market="los_angeles")
    persist_result = persist_pipedream_import_results(postgres_session, import_results)

    assert persist_result.inserted_projects == 1

    permit_identifier = postgres_session.execute(
        select(ProjectIdentifier.value).where(
            ProjectIdentifier.identifier_type == IdentifierType.PERMIT_NUMBER,
            ProjectIdentifier.value == "18010-10000-03620",
        )
    ).scalar_one()
    assert permit_identifier == "18010-10000-03620"


def test_persist_import_results_writes_pipedream_snapshot_evidence(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_evidence.xlsx",
        [
            {
                "ProjectID": "997.00001",
                "Name": "Evidence Tower",
                "Developer": "Evidence Dev",
                "Address": "9907 W Example Way",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90017",
                "CurrStatus": "Pending",
                "TotUnits": 120,
                "ProdType": "Apartment",
            }
        ],
    )

    import_results = ingest_pipedream_workbooks([workbook_path], market="los_angeles")
    persist_pipedream_import_results(postgres_session, import_results)

    project_id = postgres_session.execute(
        select(ProjectIdentifier.project_id).where(
            ProjectIdentifier.identifier_type == IdentifierType.TCG_PIPEDREAM_ID,
            ProjectIdentifier.value == "997.00001",
        )
    ).scalar_one()
    evidence_row = postgres_session.execute(
        select(Evidence).where(
            Evidence.project_id == project_id,
            Evidence.source_type == "pipedream",
            Evidence.source_record_id == "997.00001",
        )
    ).scalar_one()

    assert evidence_row.ingest_method == "seed_import"
    assert evidence_row.extracted_fields["pipeline_status"] == {
        "value": "Pending",
        "confidence": None,
    }
    assert evidence_row.extracted_fields["product_type"] == {
        "value": "Apartment",
        "confidence": None,
    }


def _count_rows(session: Session, model: type[object]) -> int:
    return session.execute(select(func.count()).select_from(model)).scalar_one()


def _build_pipedream_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "DataStorage"

    headers = _ordered_headers(rows)
    for index, header in enumerate(headers, start=1):
        column = (index * 2) - 1
        worksheet.cell(row=3, column=column, value=header)
        for row_index, row_values in enumerate(rows, start=4):
            if header in row_values:
                worksheet.cell(row=row_index, column=column, value=row_values[header])

    workbook.save(path)
    workbook.close()
    return path


def _ordered_headers(rows: list[dict[str, object]]) -> list[str]:
    preferred_order = [
        "ProjectID",
        "Address",
        "State",
        "County",
        "City",
        "CurrStatus",
        "CurrStatusDate",
        "RelP1",
    ]
    seen = set(preferred_order)
    dynamic_headers = []
    for row in rows:
        for header in row:
            if header not in seen:
                dynamic_headers.append(header)
                seen.add(header)
    return preferred_order + dynamic_headers
