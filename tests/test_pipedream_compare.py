from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.orm import Session

from tcg_pipeline.db.models import IdentifierType, PipelineStatus, Project, ProjectIdentifier
from tcg_pipeline.db.seed import ingest_pipedream_workbooks
from tcg_pipeline.evaluation.pipedream_compare import compare_pipedream_coverage


def test_compare_pipedream_coverage_reports_field_disagreements(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    project = Project(
        canonical_address="100 Pipedream Compare Ave",
        raw_addresses=["100 Pipedream Compare Ave"],
        lat=34.0501,
        lng=-118.2501,
        market="los_angeles",
        city="Los Angeles",
        state="CA",
        county="Los Angeles",
        zip="90012",
        project_name="Compare Tower",
        developer="Example Developer",
        pipeline_status=PipelineStatus.UNDER_CONSTRUCTION,
        total_units=121,
        last_evidence_date=date(2026, 6, 15),
    )
    postgres_session.add(project)
    postgres_session.flush()
    postgres_session.add(
        ProjectIdentifier(
            project_id=project.id,
            identifier_type=IdentifierType.TCG_PIPEDREAM_ID,
            value="991.00001",
            source="pipedream",
            is_primary=True,
        )
    )
    postgres_session.flush()
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_compare.xlsx",
        [
            {
                "ProjectID": "991.00001",
                "Name": "Compare Tower",
                "Address": "100 Pipedream Compare Ave",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90012",
                "Lat": 34.0501,
                "Long": -118.2501,
                "Developer": "Example Developer",
                "CurrStatus": "Approved",
                "TotUnits": 120,
            },
            {
                "ProjectID": "991.00002",
                "Address": "102 Pipedream Compare Ave",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90012",
                "CurrStatus": "Approved",
            },
        ],
    )
    import_results = ingest_pipedream_workbooks([workbook_path], market="los_angeles")

    result = compare_pipedream_coverage(
        postgres_session,
        import_results,
        market="los_angeles",
        publication_date=date(2026, 6, 15),
        compare_window_days=28,
        zip_codes=["90012"],
    )

    assert result.compared_count == 1
    assert result.projects_with_disagreements_count == 1
    assert result.field_disagreement_count == 2
    assert result.unmatched_pipedream_ids == ["991.00002"]
    comparison = result.compared_projects[0]
    assert comparison.project_id == project.id
    statuses = {field.field_name: field.status for field in comparison.fields}
    assert statuses == {
        "pipeline_status": "mismatch",
        "developer": "match",
        "total_units": "mismatch",
        "location": "match",
    }


def test_compare_pipedream_coverage_applies_evidence_window(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    project = Project(
        canonical_address="200 Old Evidence Way",
        raw_addresses=["200 Old Evidence Way"],
        market="los_angeles",
        city="Los Angeles",
        state="CA",
        county="Los Angeles",
        zip="90013",
        pipeline_status=PipelineStatus.APPROVED,
        total_units=50,
        last_evidence_date=date(2026, 1, 1),
    )
    postgres_session.add(project)
    postgres_session.flush()
    postgres_session.add(
        ProjectIdentifier(
            project_id=project.id,
            identifier_type=IdentifierType.TCG_PIPEDREAM_ID,
            value="992.00001",
            source="pipedream",
            is_primary=True,
        )
    )
    postgres_session.flush()
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_compare_window.xlsx",
        [
            {
                "ProjectID": "992.00001",
                "Address": "200 Old Evidence Way",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90013",
                "CurrStatus": "Approved",
                "TotUnits": 50,
            }
        ],
    )
    import_results = ingest_pipedream_workbooks([workbook_path], market="los_angeles")

    result = compare_pipedream_coverage(
        postgres_session,
        import_results,
        market="los_angeles",
        publication_date=date(2026, 6, 15),
        compare_window_days=28,
    )

    assert result.compared_count == 0
    assert result.excluded_evidence_window_count == 1


def test_compare_pipedream_coverage_treats_float_units_as_same_integer(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    project = Project(
        canonical_address="300 Float Units Way",
        raw_addresses=["300 Float Units Way"],
        market="los_angeles",
        city="Los Angeles",
        state="CA",
        county="Los Angeles",
        zip="90014",
        pipeline_status=PipelineStatus.APPROVED,
        total_units=120,
        last_evidence_date=date(2026, 6, 15),
    )
    postgres_session.add(project)
    postgres_session.flush()
    postgres_session.add(
        ProjectIdentifier(
            project_id=project.id,
            identifier_type=IdentifierType.TCG_PIPEDREAM_ID,
            value="993.00001",
            source="pipedream",
            is_primary=True,
        )
    )
    postgres_session.flush()
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_float_units.xlsx",
        [
            {
                "ProjectID": "993.00001",
                "Address": "300 Float Units Way",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90014",
                "CurrStatus": "Approved",
                "TotUnits": 120.0,
            }
        ],
    )
    import_results = ingest_pipedream_workbooks([workbook_path], market="los_angeles")

    result = compare_pipedream_coverage(
        postgres_session,
        import_results,
        market="los_angeles",
        publication_date=date(2026, 6, 15),
        compare_window_days=28,
    )

    comparison = result.compared_projects[0]
    units = next(field for field in comparison.fields if field.field_name == "total_units")
    assert units.status == "match"


def test_compare_pipedream_coverage_applies_location_tolerance(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    project = Project(
        canonical_address="400 Tolerance Way",
        raw_addresses=["400 Tolerance Way"],
        lat=34.0500,
        lng=-118.2500,
        market="los_angeles",
        city="Los Angeles",
        state="CA",
        county="Los Angeles",
        zip="90015",
        pipeline_status=PipelineStatus.APPROVED,
        total_units=80,
        last_evidence_date=date(2026, 6, 15),
    )
    postgres_session.add(project)
    postgres_session.flush()
    postgres_session.add(
        ProjectIdentifier(
            project_id=project.id,
            identifier_type=IdentifierType.TCG_PIPEDREAM_ID,
            value="994.00001",
            source="pipedream",
            is_primary=True,
        )
    )
    postgres_session.flush()
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_location_tolerance.xlsx",
        [
            {
                "ProjectID": "994.00001",
                "Address": "400 Tolerance Way",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90015",
                "Lat": 34.0505,
                "Long": -118.2500,
                "CurrStatus": "Approved",
                "TotUnits": 80,
            }
        ],
    )
    import_results = ingest_pipedream_workbooks([workbook_path], market="los_angeles")

    loose_result = compare_pipedream_coverage(
        postgres_session,
        import_results,
        market="los_angeles",
        publication_date=date(2026, 6, 15),
        compare_window_days=28,
        location_tolerance_meters=60.0,
    )
    tight_result = compare_pipedream_coverage(
        postgres_session,
        import_results,
        market="los_angeles",
        publication_date=date(2026, 6, 15),
        compare_window_days=28,
        location_tolerance_meters=50.0,
    )

    loose_location = next(
        field
        for field in loose_result.compared_projects[0].fields
        if field.field_name == "location"
    )
    tight_location = next(
        field
        for field in tight_result.compared_projects[0].fields
        if field.field_name == "location"
    )
    assert loose_location.status == "match"
    assert tight_location.status == "mismatch"
    assert loose_location.detail["distance_meters"] == tight_location.detail[
        "distance_meters"
    ]


def test_compare_pipedream_coverage_uses_address_fallback_without_coordinates(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    project = Project(
        canonical_address="500 Address Fallback Ave",
        raw_addresses=["500 Address Fallback Ave"],
        market="los_angeles",
        city="Los Angeles",
        state="CA",
        county="Los Angeles",
        zip="90016",
        pipeline_status=PipelineStatus.APPROVED,
        total_units=90,
        last_evidence_date=date(2026, 6, 15),
    )
    postgres_session.add(project)
    postgres_session.flush()
    postgres_session.add(
        ProjectIdentifier(
            project_id=project.id,
            identifier_type=IdentifierType.TCG_PIPEDREAM_ID,
            value="995.00001",
            source="pipedream",
            is_primary=True,
        )
    )
    postgres_session.flush()
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_location_address_fallback.xlsx",
        [
            {
                "ProjectID": "995.00001",
                "Address": "  500 address fallback ave  ",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90016",
                "CurrStatus": "Approved",
                "TotUnits": 90,
            }
        ],
    )
    import_results = ingest_pipedream_workbooks([workbook_path], market="los_angeles")

    result = compare_pipedream_coverage(
        postgres_session,
        import_results,
        market="los_angeles",
        publication_date=date(2026, 6, 15),
        compare_window_days=28,
    )

    location = next(
        field
        for field in result.compared_projects[0].fields
        if field.field_name == "location"
    )
    assert location.status == "mismatch"


def test_compare_pipedream_coverage_reports_one_sided_missing_coordinates(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    tcg_only_location = Project(
        canonical_address="600 TCG Coordinates Way",
        raw_addresses=["600 TCG Coordinates Way"],
        lat=34.0600,
        lng=-118.2600,
        market="los_angeles",
        city="Los Angeles",
        state="CA",
        county="Los Angeles",
        zip="90017",
        pipeline_status=PipelineStatus.APPROVED,
        total_units=100,
        last_evidence_date=date(2026, 6, 15),
    )
    pipedream_only_location = Project(
        canonical_address="602 Pipedream Coordinates Way",
        raw_addresses=["602 Pipedream Coordinates Way"],
        market="los_angeles",
        city="Los Angeles",
        state="CA",
        county="Los Angeles",
        zip="90017",
        pipeline_status=PipelineStatus.APPROVED,
        total_units=101,
        last_evidence_date=date(2026, 6, 15),
    )
    postgres_session.add_all([tcg_only_location, pipedream_only_location])
    postgres_session.flush()
    postgres_session.add_all(
        [
            ProjectIdentifier(
                project_id=tcg_only_location.id,
                identifier_type=IdentifierType.TCG_PIPEDREAM_ID,
                value="996.00001",
                source="pipedream",
                is_primary=True,
            ),
            ProjectIdentifier(
                project_id=pipedream_only_location.id,
                identifier_type=IdentifierType.TCG_PIPEDREAM_ID,
                value="996.00002",
                source="pipedream",
                is_primary=True,
            ),
        ]
    )
    postgres_session.flush()
    workbook_path = _build_pipedream_workbook(
        tmp_path / "pipedream_location_missing_coordinates.xlsx",
        [
            {
                "ProjectID": "996.00001",
                "Address": "600 TCG Coordinates Way",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90017",
                "CurrStatus": "Approved",
                "TotUnits": 100,
            },
            {
                "ProjectID": "996.00002",
                "Address": "602 Pipedream Coordinates Way",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90017",
                "Lat": 34.0605,
                "Long": -118.2605,
                "CurrStatus": "Approved",
                "TotUnits": 101,
            },
        ],
    )
    import_results = ingest_pipedream_workbooks([workbook_path], market="los_angeles")

    result = compare_pipedream_coverage(
        postgres_session,
        import_results,
        market="los_angeles",
        publication_date=date(2026, 6, 15),
        compare_window_days=28,
    )

    statuses = {
        comparison.pipedream_project_id: next(
            field
            for field in comparison.fields
            if field.field_name == "location"
        ).status
        for comparison in result.compared_projects
    }
    assert statuses == {
        "996.00001": "missing_pipedream",
        "996.00002": "missing_tcg",
    }


def _build_pipedream_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "DataStorage"
    headers = [
        "ProjectID",
        "Name",
        "Developer",
        "Address",
        "State",
        "County",
        "City",
        "Zip",
        "Lat",
        "Long",
        "TotUnits",
        "CurrStatus",
    ]
    for index, header in enumerate(headers, start=1):
        column = (index * 2) - 1
        worksheet.cell(row=3, column=column, value=header)
        for row_index, row_values in enumerate(rows, start=4):
            if header in row_values:
                worksheet.cell(row=row_index, column=column, value=row_values[header])
    workbook.save(path)
    workbook.close()
    return path
