from __future__ import annotations

from contextlib import contextmanager
from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import func, inspect, select
from sqlalchemy.orm import Session
from typer.testing import CliRunner

from tcg_pipeline.cli import app, resolve_all
from tcg_pipeline.db.evidence import write_evidence
from tcg_pipeline.db.models import (
    ChangeLog,
    Evidence,
    IdentifierType,
    PipelineStatus,
    Priority,
    Project,
    ProjectIdentifier,
    ProjectSourceRecord,
    ResolutionLog,
    ReviewItem,
    ReviewItemType,
    StatusHistory,
)
from tcg_pipeline.db.seed import (
    ingest_costar_workbooks,
    ingest_pipedream_workbooks,
    persist_costar_import_result,
    persist_pipedream_import_results,
)

runner = CliRunner()


def _ensure_resolution_log_metadata(postgres_session: Session) -> None:
    inspector = inspect(postgres_session.bind)
    if not inspector.has_table("resolution_log"):
        pytest.skip("Apply resolution-log migrations before running regression tests.")
    column_names = {column["name"] for column in inspector.get_columns("resolution_log")}
    if "metadata" not in column_names:
        pytest.skip("Apply the status regression metadata migration before running this test.")


def test_persist_costar_import_result_merges_into_existing_project_by_apn(
    postgres_session: Session,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    pipedream_collected_at = datetime(2026, 1, 10, 12, 0, tzinfo=UTC)
    costar_collected_at = datetime(2026, 1, 9, 12, 0, tzinfo=UTC)
    monkeypatch.setattr(
        "tcg_pipeline.db.evidence.derive_pipedream_collected_at",
        lambda project, source_record: pipedream_collected_at,
    )
    monkeypatch.setattr(
        "tcg_pipeline.db.evidence.derive_pipedream_evidence_date",
        lambda project, source_record: pipedream_collected_at.date(),
    )
    monkeypatch.setattr(
        "tcg_pipeline.db.evidence.derive_source_record_collected_at",
        lambda source_record: costar_collected_at,
    )

    pipedream_path = _build_pipedream_workbook(
        tmp_path / "pipedream_apn.xlsx",
        [
            {
                "ProjectID": "200.00001",
                "Name": "Westlake Homes",
                "Developer": "TCG Research",
                "Address": "9902 S Example Ave",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90057",
                "CurrStatus": "Pending",
                "APN": 9991001001,
            }
        ],
    )
    costar_path = _build_costar_workbook(
        tmp_path / "costar_apn.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "Property Name",
            "City",
            "State",
            "Zip",
            "County Name",
            "Parcel Number 1(Min)",
            "Constr Status",
            "Construction Begin",
            "Submarket Name",
            "Zoning",
            "Developer Name",
            "Owner Name",
            "Style",
            "RBA",
        ],
        rows=[
            {
                "PropertyID": "CST-9001",
                "Property Address": "9902 S Example Ave",
                "Property Name": "Westlake Tower",
                "City": "Los Angeles CBD",
                "State": "CA",
                "Zip": "90057-3106",
                "County Name": "Los Angeles",
                "Parcel Number 1(Min)": "9991001001",
                "Constr Status": "Under Construction",
                "Construction Begin": "December 2026",
                "Submarket Name": "Koreatown",
                "Zoning": "C2",
                "Developer Name": "Different Dev",
                "Owner Name": "Example Owner",
                "Style": "Hi-Rise",
                "RBA": 150000,
            }
        ],
    )

    pipedream_import_results = ingest_pipedream_workbooks([pipedream_path], market="los_angeles")
    persist_pipedream_import_results(postgres_session, pipedream_import_results)

    costar_import_result = ingest_costar_workbooks([costar_path], market="los_angeles")
    persist_result = persist_costar_import_result(postgres_session, costar_import_result)

    assert persist_result.inserted_projects == 0
    assert persist_result.matched_existing_projects == 1
    assert persist_result.matched_by_apn == 1
    assert persist_result.inserted_identifiers == 1
    assert persist_result.inserted_source_records == 1
    assert persist_result.inserted_status_history_entries == 1

    project = postgres_session.execute(
        select(Project).where(
            Project.canonical_address == "9902 SOUTH EXAMPLE AVENUE LOS ANGELES CA 90057"
        )
    ).scalar_one()
    # CoStar construction dates are future projections, so they do not make the row
    # artificially newer on their own. Pipedream still wins here because its snapshot
    # is newer once the projected CoStar date is capped back to collection time.
    assert project.developer == "TCG Research"
    assert project.zoning == "C2"
    assert project.owner == "Example Owner"
    assert project.costar_submarket == "Koreatown"
    assert project.style == "Hi-Rise"
    assert project.total_sf == 150000
    assert project.date_construction_start.isoformat() == "2026-12-01"

    identifier_rows = postgres_session.execute(
        select(ProjectIdentifier.identifier_type, ProjectIdentifier.value).where(
            ProjectIdentifier.project_id == project.id
        )
    ).all()
    assert (IdentifierType.COSTAR_PROPERTY_ID, "CST-9001") in identifier_rows

    source_record_ids = postgres_session.execute(
        select(ProjectSourceRecord.source_record_id).where(
            ProjectSourceRecord.project_id == project.id
        )
    ).scalars()
    assert sorted(source_record_ids) == ["200.00001", "CST-9001"]


def test_persist_costar_import_result_falls_back_to_address_match(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    pipedream_path = _build_pipedream_workbook(
        tmp_path / "pipedream_address.xlsx",
        [
            {
                "ProjectID": "201.00001",
                "Address": "8801 N Example Pl",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90028",
                "CurrStatus": "Approved",
            }
        ],
    )
    costar_path = _build_costar_workbook(
        tmp_path / "costar_address.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
            "Owner Name",
        ],
        rows=[
            {
                "PropertyID": "CST-9002",
                "Property Address": "8801 N Example Pl",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90028",
                "County Name": "Los Angeles",
                "Constr Status": "Final Planning",
                "Owner Name": "Address Match Owner",
            }
        ],
    )

    pipedream_import_results = ingest_pipedream_workbooks([pipedream_path], market="los_angeles")
    persist_pipedream_import_results(postgres_session, pipedream_import_results)

    costar_import_result = ingest_costar_workbooks([costar_path], market="los_angeles")
    persist_result = persist_costar_import_result(postgres_session, costar_import_result)

    assert persist_result.inserted_projects == 0
    assert persist_result.matched_existing_projects == 1
    assert persist_result.matched_by_address == 1
    assert persist_result.inserted_identifiers == 1


def test_persist_costar_import_result_creates_low_priority_regression_review(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    _ensure_resolution_log_metadata(postgres_session)
    project = Project(
        canonical_address="9913 SOUTH EXAMPLE AVENUE LOS ANGELES CA 90057",
        raw_addresses=["9913 S Example Ave"],
        market="los_angeles",
        city="Los Angeles",
        state="CA",
        county="Los Angeles",
        pipeline_status=PipelineStatus.UNDER_CONSTRUCTION,
    )
    postgres_session.add(project)
    postgres_session.flush()
    postgres_session.add(
        ProjectIdentifier(
            project_id=project.id,
            identifier_type=IdentifierType.COSTAR_PROPERTY_ID,
            value="CST-REGRESSION-1",
            source="costar",
            is_primary=True,
        )
    )
    write_evidence(
        postgres_session,
        project_id=project.id,
        source_name="ladbs_inspections",
        source_record_id="support-inspection-costar-seed-regression",
        raw_data={"inspection": "Frame"},
        mapped_fields={
            "status_evidence_type": "building_inspection_recorded",
            "status_evidence_date": "2026-05-01",
        },
        collected_at=datetime(2026, 5, 1, 12, 0, tzinfo=UTC),
        ingest_method="test",
    )
    postgres_session.flush()

    costar_path = _build_costar_workbook(
        tmp_path / "costar_status_regression.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
        ],
        rows=[
            {
                "PropertyID": "CST-REGRESSION-1",
                "Property Address": "9913 S Example Ave",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90057",
                "County Name": "Los Angeles",
                "Constr Status": "Final Planning",
            }
        ],
    )

    costar_import_result = ingest_costar_workbooks([costar_path], market="los_angeles")
    persist_result = persist_costar_import_result(postgres_session, costar_import_result)
    postgres_session.flush()

    assert persist_result.status_regression_review_items == 1
    assert len(persist_result.status_regression_review_item_ids) == 1
    review_item = postgres_session.execute(
        select(ReviewItem).where(
            ReviewItem.project_id == project.id,
            ReviewItem.item_type == ReviewItemType.STATUS_REGRESSION_REVIEW,
        )
    ).scalar_one()
    assert review_item.id == persist_result.status_regression_review_item_ids[0]
    assert review_item.source_run_id is None
    assert review_item.priority == Priority.LOW
    assert review_item.payload["agent_recommendation"] is None
    assert review_item.payload["system_recommendation"]["action"] == (
        "keep_current_recommended"
    )
    assert "CoStar upload from" in review_item.payload["human_summary"]
    assert "lists this project as Approved" in review_item.payload["human_summary"]
    assert (
        "LADBS inspection evidence from May 1, 2026 supports Under Construction"
        in review_item.payload["human_summary"]
    )


def test_persist_costar_import_result_is_idempotent_for_existing_property_id(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    costar_path = _build_costar_workbook(
        tmp_path / "costar_repeat.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
            "Parcel Number 1(Min)",
        ],
        rows=[
            {
                "PropertyID": "CST-9003",
                "Property Address": "8701 S Example Blvd",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90020",
                "County Name": "Los Angeles",
                "Constr Status": "Proposed",
                "Parcel Number 1(Min)": "9991001020",
            }
        ],
    )

    first_import_result = ingest_costar_workbooks([costar_path], market="los_angeles")
    second_import_result = ingest_costar_workbooks([costar_path], market="los_angeles")
    first_persist_result = persist_costar_import_result(postgres_session, first_import_result)
    second_persist_result = persist_costar_import_result(postgres_session, second_import_result)

    assert first_persist_result.inserted_projects == 1
    assert second_persist_result.inserted_projects == 0
    assert second_persist_result.matched_existing_projects == 1
    assert second_persist_result.matched_by_costar_property_id == 1
    assert second_persist_result.inserted_identifiers == 0
    assert second_persist_result.updated_source_records == 1
    assert second_persist_result.skipped_existing_status_history_entries == 1


def test_persist_costar_import_result_matches_pending_project_created_earlier_in_run(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    costar_path = _build_costar_workbook(
        tmp_path / "costar_pending_match.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
            "Parcel Number 1(Min)",
            "Owner Name",
        ],
        rows=[
            {
                "PropertyID": "CST-9101",
                "Property Address": "8500 W Example Ave",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90036",
                "County Name": "Los Angeles",
                "Constr Status": "Proposed",
                "Parcel Number 1(Min)": "9991002001",
            },
            {
                "PropertyID": "CST-9102",
                "Property Address": "8500 W Example Ave",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90036",
                "County Name": "Los Angeles",
                "Constr Status": "Final Planning",
                "Parcel Number 1(Min)": "9991002001",
                "Owner Name": "Pending Match Owner",
            },
        ],
    )

    import_result = ingest_costar_workbooks([costar_path], market="los_angeles")
    persist_result = persist_costar_import_result(postgres_session, import_result)

    assert persist_result.inserted_projects == 1
    assert persist_result.matched_existing_projects == 1
    assert persist_result.matched_by_apn == 1
    assert persist_result.inserted_identifiers == 3
    assert persist_result.skipped_existing_identifiers == 1
    assert persist_result.inserted_source_records == 2

    project = postgres_session.execute(
        select(Project).where(
            Project.canonical_address == "8500 WEST EXAMPLE AVENUE LOS ANGELES CA 90036"
        )
    ).scalar_one()
    assert project.owner == "Pending Match Owner"


def test_seed_costar_command_reports_merge_counts(
    monkeypatch: pytest.MonkeyPatch,
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    pipedream_path = _build_pipedream_workbook(
        tmp_path / "pipedream_cli.xlsx",
        [
            {
                "ProjectID": "202.00001",
                "Address": "8601-8603 E Example St",
                "State": "CA",
                "County": "Los Angeles",
                "City": "Los Angeles",
                "Zip": "90013",
                "CurrStatus": "Pending",
            }
        ],
    )
    costar_path = _build_costar_workbook(
        tmp_path / "costar_cli.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
            "Owner Name",
        ],
        rows=[
            {
                "PropertyID": "CST-9004",
                "Property Address": "8601-8603 E Example St",
                "City": "Downtown Los Angeles",
                "State": "CA",
                "Zip": "90013",
                "County Name": "Los Angeles",
                "Constr Status": "Proposed",
                "Owner Name": "CLI Owner",
            }
        ],
    )

    pipedream_import_results = ingest_pipedream_workbooks([pipedream_path], market="los_angeles")
    persist_pipedream_import_results(postgres_session, pipedream_import_results)

    @contextmanager
    def fake_session_factory() -> Session:
        yield postgres_session

    monkeypatch.setattr("tcg_pipeline.cli.get_session_factory", lambda: fake_session_factory)

    result = runner.invoke(
        app,
        ["seed-costar", str(costar_path), "--market", "los_angeles"],
    )

    assert result.exit_code == 0
    assert "Imported projects: 1" in result.stdout
    assert "Matched existing projects: 1" in result.stdout
    assert "Matched by address: 1" in result.stdout
    assert "Inserted source records: 1" in result.stdout


def test_seed_costar_with_defer_resolve_skips_resolution_writes(
    monkeypatch: pytest.MonkeyPatch,
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    costar_path = _build_costar_workbook(
        tmp_path / "costar_defer_resolve.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
            "Number Of Units",
        ],
        rows=[
            {
                "PropertyID": "CST-DEFER-1",
                "Property Address": "8702 W Deferred Way",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90018",
                "County Name": "Los Angeles",
                "Constr Status": "Under Construction",
                "Number Of Units": 88,
            }
        ],
    )

    @contextmanager
    def fake_session_factory() -> Session:
        yield postgres_session

    monkeypatch.setattr("tcg_pipeline.cli.get_session_factory", lambda: fake_session_factory)

    result = runner.invoke(
        app,
        ["seed-costar", str(costar_path), "--market", "los_angeles", "--defer-resolve"],
    )

    assert result.exit_code == 0
    assert "Deferred inline resolve: True" in result.stdout
    project_id = postgres_session.execute(
        select(ProjectIdentifier.project_id).where(
            ProjectIdentifier.identifier_type == IdentifierType.COSTAR_PROPERTY_ID,
            ProjectIdentifier.value == "CST-DEFER-1",
        )
    ).scalar_one()
    project = postgres_session.get(Project, project_id)

    assert project is not None
    assert project.last_evidence_date is None
    assert _count_rows(postgres_session, Evidence) == 1
    assert _count_rows(postgres_session, ProjectSourceRecord) == 1
    assert _count_rows(postgres_session, StatusHistory) == 1
    assert _count_rows(postgres_session, ResolutionLog) == 0
    assert _count_rows(postgres_session, ChangeLog) == 0


def test_seed_costar_without_flag_resolves_inline(
    monkeypatch: pytest.MonkeyPatch,
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    costar_path = _build_costar_workbook(
        tmp_path / "costar_inline_resolve.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
            "Number Of Units",
        ],
        rows=[
            {
                "PropertyID": "CST-INLINE-1",
                "Property Address": "8703 W Inline Way",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90018",
                "County Name": "Los Angeles",
                "Constr Status": "Under Construction",
                "Number Of Units": 90,
            }
        ],
    )

    @contextmanager
    def fake_session_factory() -> Session:
        yield postgres_session

    monkeypatch.setattr("tcg_pipeline.cli.get_session_factory", lambda: fake_session_factory)

    result = runner.invoke(
        app,
        ["seed-costar", str(costar_path), "--market", "los_angeles"],
    )

    assert result.exit_code == 0
    assert "Deferred inline resolve: False" in result.stdout
    project_id = postgres_session.execute(
        select(ProjectIdentifier.project_id).where(
            ProjectIdentifier.identifier_type == IdentifierType.COSTAR_PROPERTY_ID,
            ProjectIdentifier.value == "CST-INLINE-1",
        )
    ).scalar_one()
    project = postgres_session.get(Project, project_id)

    assert project is not None
    assert project.last_evidence_date is not None
    assert _count_rows(postgres_session, ResolutionLog) > 0


def test_resolve_all_apply_completes_deferred_costar_seed_state(
    monkeypatch: pytest.MonkeyPatch,
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    costar_path = _build_costar_workbook(
        tmp_path / "costar_deferred_resolve_all.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
            "Number Of Units",
        ],
        rows=[
            {
                "PropertyID": "CST-RESOLVE-ALL-1",
                "Property Address": "8704 W Resolve All Way",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90018",
                "County Name": "Los Angeles",
                "Constr Status": "Under Construction",
                "Number Of Units": 92,
            }
        ],
    )

    @contextmanager
    def fake_session_factory() -> Session:
        yield postgres_session

    monkeypatch.setattr("tcg_pipeline.cli.get_session_factory", lambda: fake_session_factory)

    seed_result = runner.invoke(
        app,
        ["seed-costar", str(costar_path), "--market", "los_angeles", "--defer-resolve"],
    )
    assert seed_result.exit_code == 0

    project_id = postgres_session.execute(
        select(ProjectIdentifier.project_id).where(
            ProjectIdentifier.identifier_type == IdentifierType.COSTAR_PROPERTY_ID,
            ProjectIdentifier.value == "CST-RESOLVE-ALL-1",
        )
    ).scalar_one()
    project = postgres_session.get(Project, project_id)
    assert project is not None
    assert project.last_evidence_date is None
    assert _count_rows(postgres_session, ResolutionLog) == 0
    assert _count_rows(postgres_session, StatusHistory) == 1

    resolve_all(market="los_angeles", apply=True, clear_log=True, limit=1, batch_size=1)
    postgres_session.expire_all()
    project = postgres_session.get(Project, project_id)

    assert project is not None
    assert project.last_evidence_date is not None
    assert project.confidence_reason is not None
    assert project.likelihood is not None
    assert _count_rows(postgres_session, ResolutionLog) > 0
    assert _count_rows(postgres_session, StatusHistory) == 1


def test_persist_costar_import_result_writes_costar_evidence(
    postgres_session: Session,
    tmp_path: Path,
) -> None:
    costar_path = _build_costar_workbook(
        tmp_path / "costar_evidence.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "Property Name",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
            "Developer Name",
            "Number Of Units",
        ],
        rows=[
            {
                "PropertyID": "CST-9901",
                "Property Address": "8707 W Example Terrace",
                "Property Name": "Evidence Plaza",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90018",
                "County Name": "Los Angeles",
                "Constr Status": "Proposed",
                "Developer Name": "Costar Dev",
                "Number Of Units": 42,
            }
        ],
    )

    import_result = ingest_costar_workbooks([costar_path], market="los_angeles")
    persist_costar_import_result(postgres_session, import_result)

    project_id = postgres_session.execute(
        select(ProjectIdentifier.project_id).where(
            ProjectIdentifier.identifier_type == IdentifierType.COSTAR_PROPERTY_ID,
            ProjectIdentifier.value == "CST-9901",
        )
    ).scalar_one()
    evidence_row = postgres_session.execute(
        select(Evidence).where(
            Evidence.project_id == project_id,
            Evidence.source_type == "costar",
            Evidence.source_record_id == "CST-9901",
        )
    ).scalar_one()

    assert evidence_row.ingest_method == "seed_import"
    assert evidence_row.extracted_fields["pipeline_status"] == {
        "value": "Proposed",
        "confidence": None,
    }
    assert evidence_row.extracted_fields["developer"] == {
        "value": "Costar Dev",
        "confidence": None,
    }


def _count_rows(session: Session, model: type[object]) -> int:
    return session.execute(select(func.count()).select_from(model)).scalar_one()


def _build_pipedream_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "DataStorage"

    headers = _ordered_pipedream_headers(rows)
    for index, header in enumerate(headers, start=1):
        column = (index * 2) - 1
        worksheet.cell(row=3, column=column, value=header)
        for row_index, row_values in enumerate(rows, start=4):
            if header in row_values:
                worksheet.cell(row=row_index, column=column, value=row_values[header])

    workbook.save(path)
    workbook.close()
    return path


def _ordered_pipedream_headers(rows: list[dict[str, object]]) -> list[str]:
    preferred_order = [
        "ProjectID",
        "Name",
        "Developer",
        "Address",
        "State",
        "County",
        "City",
        "Zip",
        "CurrStatus",
        "APN",
    ]
    seen = set(preferred_order)
    dynamic_headers = []
    for row in rows:
        for header in row:
            if header not in seen:
                dynamic_headers.append(header)
                seen.add(header)
    return preferred_order + dynamic_headers


def _build_costar_workbook(
    path: Path,
    *,
    headers: list[str],
    rows: list[dict[str, object]],
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Export041526"

    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column_index, value=header)
        for row_index, row_values in enumerate(rows, start=2):
            if header in row_values:
                worksheet.cell(row=row_index, column=column_index, value=row_values[header])

    workbook.save(path)
    workbook.close()
    return path
