from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from tcg_pipeline.cli import app
from tcg_pipeline.db.models import (
    AgeRestriction,
    IdentifierType,
    PipelineStatus,
    ProductType,
    RentOrSale,
)
from tcg_pipeline.ingesters.costar import CoStarIngester

runner = CliRunner()


def test_ingest_costar_multifamily_workbook_maps_by_header_name(tmp_path: Path) -> None:
    workbook_path = _build_costar_workbook(
        tmp_path / "costar_mf.xlsx",
        headers=[
            "Developer Name",
            "Property Address",
            "PropertyID",
            "Constr Status",
            "City",
            "State",
            "Zip",
            "County Name",
            "Submarket Name",
            "Latitude",
            "Longitude",
            "Parcel Number 1(Min)",
            "Parcel Number 2(Max)",
            "Property Type",
            "Secondary Type",
            "Property Name",
            "Number Of Units",
            "RBA",
            "Number Of Stories",
            "Style",
            "Land Area (AC)",
            "Number Of Parking Spaces",
            "Zoning",
            "% Studios",
            "% 1-Bed",
            "% 2-Bed",
            "% 3-Bed",
            "% 4-Bed",
            "Rent Type",
            "Affordable Type",
            "Market Segment",
            "Construction Begin",
            "Year Built",
            "Month Built",
            "Owner Name",
            "True Owner Name",
            "Architect Name",
        ],
        rows=[
            {
                "PropertyID": "CST-1001",
                "Property Address": "602 S Westlake Ave",
                "Property Name": "Westlake Tower",
                "City": "Los Angeles CBD",
                "State": "CA",
                "Zip": "90057-3106",
                "County Name": "Los Angeles",
                "Submarket Name": "Koreatown",
                "Latitude": 34.0601,
                "Longitude": -118.2722,
                "Parcel Number 1(Min)": "5143001001",
                "Parcel Number 2(Max)": "5143001002",
                "Property Type": "Multifamily",
                "Secondary Type": "Apartments",
                "Number Of Units": 200,
                "RBA": 150000,
                "Number Of Stories": 20,
                "Style": "Hi-Rise",
                "Land Area (AC)": 1.2,
                "Number Of Parking Spaces": 250,
                "Zoning": "C2",
                "% Studios": 10,
                "% 1-Bed": 40,
                "% 2-Bed": 35,
                "% 3-Bed": 10,
                "% 4-Bed": 5,
                "Rent Type": "Market/Affordable",
                "Affordable Type": "Rent Restricted",
                "Market Segment": "Senior",
                "Constr Status": "Final Planning",
                "Construction Begin": "December 2026",
                "Year Built": 2028,
                "Month Built": "March",
                "Developer Name": "Example Dev",
                "Owner Name": "Example Owner",
                "True Owner Name": "Example Holdco",
                "Architect Name": "Example Architect",
            }
        ],
    )

    result = CoStarIngester(market="los_angeles").ingest_workbook(workbook_path)

    assert result.imported_count == 1
    assert result.issue_count == 0
    record = result.project_records[0]
    project = record.project

    assert record.property_id == "CST-1001"
    assert project.project_name == "Westlake Tower"
    assert project.canonical_address == "602 SOUTH WESTLAKE AVENUE LOS ANGELES CA 90057"
    assert project.city == "LOS ANGELES"
    assert project.pipeline_status == PipelineStatus.APPROVED
    assert project.date_construction_start.isoformat() == "2026-12-01"
    assert project.date_delivery.isoformat() == "2028-03-01"
    assert project.total_units == 200
    assert project.total_sf == 150000
    assert project.style == "Hi-Rise"
    assert project.zoning == "C2"
    assert project.costar_submarket == "Koreatown"
    assert project.rent_or_sale == RentOrSale.RENTAL
    assert project.product_type == ProductType.APARTMENT
    assert project.age_restriction == AgeRestriction.SENIOR
    assert project.pct_studio == 0.10
    assert project.pct_1bed == 0.40
    assert project.pct_2bed == 0.35
    assert project.pct_other_bed == 0.15

    identifier_types = {identifier.identifier_type for identifier in record.identifiers}
    apn_identifier_count = [
        identifier.identifier_type for identifier in record.identifiers
    ].count(IdentifierType.APN)
    assert IdentifierType.COSTAR_PROPERTY_ID in identifier_types
    assert apn_identifier_count == 2


def test_ingest_costar_dedupes_duplicate_apn_fields(tmp_path: Path) -> None:
    workbook_path = _build_costar_workbook(
        tmp_path / "costar_duplicate_apn.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
            "Parcel Number 1(Min)",
            "Parcel Number 2(Max)",
        ],
        rows=[
            {
                "PropertyID": "CST-1002",
                "Property Address": "549 S Harvard Blvd",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90020",
                "County Name": "Los Angeles",
                "Constr Status": "Proposed",
                "Parcel Number 1(Min)": "5078001020",
                "Parcel Number 2(Max)": "5078001020",
            }
        ],
    )

    result = CoStarIngester(market="los_angeles").ingest_workbook(workbook_path)

    assert result.imported_count == 1
    record = result.project_records[0]
    apn_identifier_count = [
        identifier.identifier_type for identifier in record.identifiers
    ].count(IdentifierType.APN)
    assert apn_identifier_count == 1


def test_ingest_costar_non_mf_workbook_and_diagnostics(tmp_path: Path) -> None:
    workbook_path = _build_costar_workbook(
        tmp_path / "costar_non_mf.xlsx",
        headers=[
            "PropertyID",
            "Property Type",
            "Property Address",
            "Property Name",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
            "Construction Begin",
            "Year Built",
            "Month Built",
            "RBA",
            "Number Of Stories",
            "Market Segment",
            "Submarket Name",
        ],
        rows=[
            {
                "PropertyID": "OFF-2001",
                "Property Type": "Office",
                "Property Address": "W 3rd St",
                "Property Name": "Creative Offices",
                "City": "Hollywood",
                "State": "CA",
                "Zip": "90036",
                "County Name": "Los Angeles",
                "Constr Status": "Deferred",
                "Construction Begin": "Spring 2026",
                "Year Built": 2027,
                "Month Built": 11,
                "RBA": 45000,
                "Number Of Stories": 4,
                "Market Segment": "All",
                "Submarket Name": "Hollywood",
            }
        ],
    )

    result = CoStarIngester(market="los_angeles").ingest_workbook(workbook_path)

    assert result.imported_count == 1
    assert result.issue_counts == {"invalid_date": 1}
    record = result.project_records[0]
    project = record.project

    assert project.pipeline_status == PipelineStatus.STALLED
    assert project.date_construction_start is None
    assert project.date_delivery.isoformat() == "2027-11-01"
    assert project.city == "LOS ANGELES"
    assert project.property_type == "Office"
    assert project.product_type == ProductType.UNKNOWN
    assert project.rent_or_sale == RentOrSale.UNKNOWN
    assert project.age_restriction == AgeRestriction.NON_AGE_RESTRICTED
    assert project.total_sf == 45000


def test_ingest_costar_workbooks_dedupes_property_ids_across_files(tmp_path: Path) -> None:
    workbook_one = _build_costar_workbook(
        tmp_path / "costar_one.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
        ],
        rows=[
            {
                "PropertyID": "CST-3001",
                "Property Address": "549 S Harvard Blvd",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90020",
                "County Name": "Los Angeles",
                "Constr Status": "Proposed",
            }
        ],
    )
    workbook_two = _build_costar_workbook(
        tmp_path / "costar_two.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
        ],
        rows=[
            {
                "PropertyID": "CST-3001",
                "Property Address": "1718 N Las Palmas Ave",
                "City": "Los Angeles",
                "State": "CA",
                "Zip": "90028",
                "County Name": "Los Angeles",
                "Constr Status": "Under Construction",
            }
        ],
    )

    result = CoStarIngester(market="los_angeles").ingest_workbooks([workbook_one, workbook_two])

    assert result.imported_count == 1
    assert result.duplicate_count == 1
    assert result.duplicate_property_ids == ["CST-3001"]
    assert (
        result.project_records[0].project.canonical_address
        == "549 SOUTH HARVARD BOULEVARD LOS ANGELES CA 90020"
    )


def test_preview_costar_command_reports_counts(tmp_path: Path) -> None:
    _build_costar_workbook(
        tmp_path / "costar_cli.xlsx",
        headers=[
            "PropertyID",
            "Property Address",
            "City",
            "State",
            "Zip",
            "County Name",
            "Constr Status",
        ],
        rows=[
            {
                "PropertyID": "CST-4001",
                "Property Address": "407-413 E 5th St",
                "City": "Downtown Los Angeles",
                "State": "CA",
                "Zip": "90013",
                "County Name": "Los Angeles",
                "Constr Status": "Proposed",
            }
        ],
    )

    result = runner.invoke(
        app,
        ["preview-costar", str(tmp_path), "--market", "los_angeles"],
    )

    assert result.exit_code == 0
    assert "Workbooks: 1" in result.stdout
    assert "Imported projects: 1" in result.stdout
    assert "Duplicate property ids: 0" in result.stdout
    assert "Issues: 0" in result.stdout


def _build_costar_workbook(
    path: Path,
    *,
    headers: list[str],
    rows: list[dict[str, object]],
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Export041526"

    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column_index, value=header)
        for row_index, row_values in enumerate(rows, start=2):
            if header in row_values:
                worksheet.cell(row=row_index, column=column_index, value=row_values[header])

    workbook.save(path)
    workbook.close()
    return path
