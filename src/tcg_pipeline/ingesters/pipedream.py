from __future__ import annotations

import logging
import re
from collections import Counter
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from tcg_pipeline.db.models import (
    AgeRestriction,
    DismissedRecord,
    DismissReason,
    IdentifierType,
    PipelineStatus,
    ProductType,
    Project,
    ProjectIdentifier,
    ProjectNote,
    ProjectSourceRecord,
    RelationshipType,
    RentOrSale,
    StatusConfidence,
    StatusHistory,
)
from tcg_pipeline.ingesters._common import (
    build_location as _build_location,
)
from tcg_pipeline.ingesters._common import (
    clean_identifier_text as _clean_identifier_text,
)
from tcg_pipeline.ingesters._common import (
    clean_text as _clean_text,
)
from tcg_pipeline.ingesters._common import (
    dedupe_strings as _dedupe_strings,
)
from tcg_pipeline.ingesters._common import (
    determine_geocode_confidence as _determine_geocode_confidence,
)
from tcg_pipeline.ingesters._common import (
    display_value as _display_value,
)
from tcg_pipeline.ingesters._common import (
    parse_float as _parse_float,
)
from tcg_pipeline.ingesters._common import (
    parse_int as _parse_int,
)
from tcg_pipeline.ingesters._common import (
    row_has_values as _row_has_values,
)
from tcg_pipeline.ingesters._common import (
    serialize_json_value as _serialize_json_value,
)
from tcg_pipeline.matching.normalizer import (
    normalize_address,
    normalize_city,
    normalize_postal_code,
)
from tcg_pipeline.permit_numbers import extract_ladbs_pcis_permit_numbers

logger = logging.getLogger(__name__)

HEADER_ROW_INDEX = 3
DATA_START_ROW_INDEX = 4
DATA_STORAGE_TAB_NAME = "DataStorage"
PIPEDREAM_SOURCE_NAME = "pipedream"
PIPEDREAM_CREATED_BY = "pipedream_import"
HEADER_COLUMNS = (
    "ProjectID",
    "Name",
    "Developer",
    "Address",
    "State",
    "County",
    "City",
    "Zip",
    "Region",
    "Lat",
    "Long",
    "RentFS",
    "MRUnits",
    "AffUnits",
    # Current workbooks may not expose this yet; keep the ingester forward-compatible.
    "WorkforceUnits",
    "TotUnits",
    "Acres",
    "RetailSF",
    "OfficeSF",
    "HKeys",
    "ProdType",
    "Elevation",
    "Senior",
    "PercS",
    "Perc1B",
    "Perc2B",
    "PercOther",
    "CurrStatus",
    "CurrStatusDate",
    "Jurisdiction",
    "RefNum",
    "APN",
    "Plan1Name",
    "Plan1City",
    "Plan1Email",
    "Plan1Phone",
    "Plan2Name",
    "Plan2City",
    "Plan2Email",
    "Plan2Phone",
    "Notes",
    "Site1",
    "Site2",
    "Site3",
    "Site4",
    "PersonalNotes",
    "ChangeNotes",
    "PStat1",
    "PStatDate1",
    "PStat2",
    "PStatDate2",
    "PStat3",
    "PStatDate3",
    "PStat4",
    "PStatDate4",
    "PStat5",
    "PStatDate5",
    "PStat6",
    "PStatDate6",
    "PrevName1",
    "PrevName2",
    "CorrP",
    "PCPart",
    "RelP1",
    "RelP2",
    "RelP3",
    "RelP4",
    "RelP5",
    "RelP6",
    "DeliveryDate",
    "Editor",
    "EditDate",
)
PROJECT_ID_FIELDS = {
    "ProjectID",
    "CorrP",
    "PCPart",
    "RelP1",
    "RelP2",
    "RelP3",
    "RelP4",
    "RelP5",
    "RelP6",
}
STATUS_FIELD_PAIRS = (
    ("PStat6", "PStatDate6"),
    ("PStat5", "PStatDate5"),
    ("PStat4", "PStatDate4"),
    ("PStat3", "PStatDate3"),
    ("PStat2", "PStatDate2"),
    ("PStat1", "PStatDate1"),
)
SOURCE_URL_FIELDS = ("Site1", "Site2", "Site3", "Site4")
PREVIOUS_NAME_FIELDS = ("PrevName1", "PrevName2")
RELATIONSHIP_FIELD_MAP = {
    "CorrP": RelationshipType.DUPLICATE,
    "PCPart": RelationshipType.COUNTERPART,
    "RelP1": RelationshipType.PHASE,
    "RelP2": RelationshipType.PHASE,
    "RelP3": RelationshipType.PHASE,
    "RelP4": RelationshipType.PHASE,
    "RelP5": RelationshipType.PHASE,
    "RelP6": RelationshipType.PHASE,
}
DELETE_STATUS_REASON_MAP = {
    "Delete - Duplicate": DismissReason.DUPLICATE,
    "Delete - Outside Market Area": DismissReason.OUTSIDE_MARKET,
    "Delete - Not Residential": DismissReason.NOT_RESIDENTIAL,
}
PIPELINE_STATUS_MAP = {
    PipelineStatus.CONCEPTUAL.value: PipelineStatus.CONCEPTUAL,
    PipelineStatus.PROPOSED.value: PipelineStatus.PROPOSED,
    PipelineStatus.PENDING.value: PipelineStatus.PENDING,
    PipelineStatus.APPROVED.value: PipelineStatus.APPROVED,
    PipelineStatus.UNDER_CONSTRUCTION.value: PipelineStatus.UNDER_CONSTRUCTION,
    PipelineStatus.PRE_LEASING_PRE_SELLING.value: PipelineStatus.PRE_LEASING_PRE_SELLING,
    PipelineStatus.COMPLETE.value: PipelineStatus.COMPLETE,
    PipelineStatus.STALLED.value: PipelineStatus.STALLED,
    PipelineStatus.INACTIVE.value: PipelineStatus.INACTIVE,
    "Delete - Duplicate": PipelineStatus.DELETE_DUPLICATE,
    "Delete - Outside Market Area": PipelineStatus.DELETE_OUTSIDE_MARKET_AREA,
    "Delete - Not Residential": PipelineStatus.DELETE_NOT_RESIDENTIAL,
}
RENT_OR_SALE_MAP = {
    RentOrSale.RENTAL.value: RentOrSale.RENTAL,
    RentOrSale.FOR_SALE.value: RentOrSale.FOR_SALE,
    RentOrSale.BOTH.value: RentOrSale.BOTH,
    "Both (Rental & FS)": RentOrSale.BOTH,
    RentOrSale.UNKNOWN.value: RentOrSale.UNKNOWN,
}
PRODUCT_TYPE_MAP = {
    ProductType.APARTMENT.value: ProductType.APARTMENT,
    ProductType.CONDO.value: ProductType.CONDO,
    ProductType.SINGLE_FAMILY.value: ProductType.SINGLE_FAMILY,
    "Single Family": ProductType.SINGLE_FAMILY,
    ProductType.TOWNHOME.value: ProductType.TOWNHOME,
    ProductType.MICRO_CO_LIVING.value: ProductType.MICRO_CO_LIVING,
    ProductType.OTHER.value: ProductType.OTHER,
    ProductType.UNKNOWN.value: ProductType.UNKNOWN,
}
AGE_RESTRICTION_MAP = {
    AgeRestriction.NON_AGE_RESTRICTED.value: AgeRestriction.NON_AGE_RESTRICTED,
    AgeRestriction.SENIOR.value: AgeRestriction.SENIOR,
    AgeRestriction.STUDENT.value: AgeRestriction.STUDENT,
    AgeRestriction.UNKNOWN.value: AgeRestriction.UNKNOWN,
}
ADDRESS_LIKE_RE = re.compile(r"^\d+[\w\s.#/-]*$")
ADDRESS_HINT_TOKENS = (
    "ST",
    "STREET",
    "AVE",
    "AVENUE",
    "BLVD",
    "BOULEVARD",
    "RD",
    "ROAD",
    "DR",
    "DRIVE",
    "LN",
    "LANE",
    "CT",
    "COURT",
    "PL",
    "PLACE",
    "WAY",
    "PKWY",
    "PARKWAY",
    "TER",
    "TERRACE",
)
ENTITY_NAME_TOKENS = (
    " LLC",
    " LP",
    " INC",
    " CORP",
    " CORPORATION",
    " HOLDINGS",
    " INVESTMENTS",
    " PARTNERS",
    " GROUP",
    " FUND",
    " COMPANY",
    " CO.",
)


@dataclass(slots=True)
class PipedreamImportIssue:
    issue_type: str
    row_number: int
    field_name: str
    raw_value: str | None
    message: str


@dataclass(slots=True)
class StagedProjectRelationship:
    project_identifier_value: str
    related_project_identifier_value: str
    relationship_type: RelationshipType
    source_field: str
    notes: str | None = None


@dataclass(slots=True)
class PipedreamProjectRecord:
    row_number: int
    project_identifier_value: str
    project: Project
    project_notes: list[ProjectNote]
    identifiers: list[ProjectIdentifier]
    status_history: list[StatusHistory]
    source_record: ProjectSourceRecord


@dataclass(slots=True)
class PipedreamImportResult:
    source_path: Path
    project_records: list[PipedreamProjectRecord] = field(default_factory=list)
    dismissed_records: list[DismissedRecord] = field(default_factory=list)
    staged_relationships: list[StagedProjectRelationship] = field(default_factory=list)
    skipped_project_ids: list[str] = field(default_factory=list)
    issues: list[PipedreamImportIssue] = field(default_factory=list)
    missing_project_id_rows: int = 0

    @property
    def imported_count(self) -> int:
        return len(self.project_records)

    @property
    def dismissed_count(self) -> int:
        return len(self.dismissed_records)

    @property
    def issue_count(self) -> int:
        return len(self.issues)

    @property
    def issue_counts(self) -> dict[str, int]:
        return dict(Counter(issue.issue_type for issue in self.issues))

    def add_issue(
        self,
        *,
        issue_type: str,
        row_number: int,
        field_name: str,
        raw_value: Any,
        message: str,
    ) -> None:
        issue = PipedreamImportIssue(
            issue_type=issue_type,
            row_number=row_number,
            field_name=field_name,
            raw_value=_display_value(raw_value),
            message=message,
        )
        if issue in self.issues:
            return
        self.issues.append(issue)
        logger.warning(
            "%s row %s field %s: %s (raw=%r)",
            self.source_path.name,
            row_number,
            field_name,
            message,
            issue.raw_value,
        )


class PipedreamIngester:
    def __init__(
        self,
        *,
        market: str,
        source_name: str = PIPEDREAM_SOURCE_NAME,
        allowed_cities: Iterable[str] | None = None,
    ) -> None:
        self.market = market
        self.source_name = source_name
        self.allowed_cities = {
            normalized
            for city in (allowed_cities or [])
            if (normalized := normalize_city(city, market=market))
        }

    def ingest_workbook(self, workbook_path: str | Path) -> PipedreamImportResult:
        path = Path(workbook_path)
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            if DATA_STORAGE_TAB_NAME not in workbook.sheetnames:
                raise ValueError(
                    f"{path.name} is missing the required '{DATA_STORAGE_TAB_NAME}' worksheet."
                )

            worksheet = workbook[DATA_STORAGE_TAB_NAME]
            column_to_header = _extract_headers(worksheet)
            imported_at = datetime.now(UTC)
            result = PipedreamImportResult(source_path=path)

            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=DATA_START_ROW_INDEX),
                start=DATA_START_ROW_INDEX,
            ):
                payload = _build_row_payload(
                    row,
                    column_to_header,
                    result=result,
                    row_number=row_number,
                )
                project_identifier = _clean_project_identifier(
                    payload.get("ProjectID"),
                    result=result,
                    row_number=row_number,
                    field_name="ProjectID",
                )
                if not project_identifier:
                    if _row_has_values(payload):
                        result.missing_project_id_rows += 1
                        logger.warning(
                            "%s row %s skipped: missing ProjectID",
                            path.name,
                            row_number,
                        )
                    continue

                if self.allowed_cities:
                    normalized_city = normalize_city(payload.get("City"), market=self.market)
                    if normalized_city not in self.allowed_cities:
                        result.skipped_project_ids.append(project_identifier)
                        continue

                raw_status = _clean_text(payload.get("CurrStatus"))
                result.staged_relationships.extend(
                    _build_relationships(
                        payload,
                        project_identifier,
                        result=result,
                        row_number=row_number,
                    )
                )

                if raw_status in DELETE_STATUS_REASON_MAP:
                    result.dismissed_records.append(
                        _build_dismissed_record(
                            payload=payload,
                            project_identifier=project_identifier,
                            reason=DELETE_STATUS_REASON_MAP[raw_status],
                            source_name=self.source_name,
                            market=self.market,
                        )
                    )
                    continue

                project_record = _build_project_record(
                    payload=payload,
                    row_number=row_number,
                    project_identifier=project_identifier,
                    market=self.market,
                    source_name=self.source_name,
                    imported_at=imported_at,
                    result=result,
                )
                result.project_records.append(project_record)

            return result
        finally:
            workbook.close()


def _extract_headers(worksheet: Worksheet) -> dict[int, str]:
    headers: dict[int, str] = {}
    for cell in worksheet[HEADER_ROW_INDEX]:
        value = _clean_text(cell.value)
        if value:
            headers[cell.column] = value
    return headers


def _build_row_payload(
    row: tuple[Any, ...],
    column_to_header: dict[int, str],
    *,
    result: PipedreamImportResult,
    row_number: int,
) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for column_index, header in column_to_header.items():
        if column_index > len(row):
            continue
        cell_value = row[column_index - 1].value
        if header in PROJECT_ID_FIELDS:
            payload[header] = _clean_project_identifier(
                cell_value,
                result=result,
                row_number=row_number,
                field_name=header,
            )
        else:
            payload[header] = cell_value
    return payload


def _build_project_record(
    *,
    payload: dict[str, Any],
    row_number: int,
    project_identifier: str,
    market: str,
    source_name: str,
    imported_at: datetime,
    result: PipedreamImportResult,
) -> PipedreamProjectRecord:
    address_value = _select_address_value(payload)
    normalized_address = normalize_address(
        address_value or "",
        city=_clean_text(payload.get("City")),
        state=_clean_text(payload.get("State")),
        postal_code=_clean_text(payload.get("Zip")),
        market=market,
    )
    project = Project(
        canonical_address=normalized_address.canonical_address or normalized_address.raw_address,
        raw_addresses=_dedupe_strings([address_value, _clean_text(payload.get("Address"))]),
        lat=_parse_float(payload.get("Lat")),
        lng=_parse_float(payload.get("Long")),
        location=_build_location(payload.get("Lat"), payload.get("Long")),
        geocode_confidence=_determine_geocode_confidence(
            payload.get("Lat"),
            payload.get("Long"),
        ),
        market=market,
        city=normalized_address.city or _clean_text(payload.get("City")) or "",
        state=normalized_address.state or _clean_text(payload.get("State")) or "",
        county=_clean_text(payload.get("County")) or "",
        zip=normalize_postal_code(payload.get("Zip")),
        tcg_region=_clean_text(payload.get("Region")),
        jurisdiction=_clean_text(payload.get("Jurisdiction")),
        project_name=_clean_text(payload.get("Name")),
        previous_names=_dedupe_strings(
            _clean_text(payload.get(field)) for field in PREVIOUS_NAME_FIELDS
        ),
        developer=_clean_text(payload.get("Developer")),
        rent_or_sale=_parse_rent_or_sale(
            payload.get("RentFS"),
            result=result,
            row_number=row_number,
            field_name="RentFS",
        ),
        product_type=_parse_product_type(
            payload.get("ProdType"),
            result=result,
            row_number=row_number,
            field_name="ProdType",
        ),
        age_restriction=_parse_age_restriction(
            payload.get("Senior"),
            result=result,
            row_number=row_number,
            field_name="Senior",
        ),
        stories=_parse_int(payload.get("Elevation")),
        total_units=_parse_int(payload.get("TotUnits")),
        market_rate_units=_parse_int(payload.get("MRUnits")),
        affordable_units=_parse_int(payload.get("AffUnits")),
        workforce_units=_parse_int(payload.get("WorkforceUnits")),
        pct_studio=_parse_float(payload.get("PercS")),
        pct_1bed=_parse_float(payload.get("Perc1B")),
        pct_2bed=_parse_float(payload.get("Perc2B")),
        pct_other_bed=_parse_float(payload.get("PercOther")),
        acres=_parse_float(payload.get("Acres")),
        retail_sf=_parse_int(payload.get("RetailSF")),
        office_sf=_parse_int(payload.get("OfficeSF")),
        hotel_keys=_parse_int(payload.get("HKeys")),
        pipeline_status=_parse_pipeline_status(
            payload.get("CurrStatus"),
            result=result,
            row_number=row_number,
            field_name="CurrStatus",
        ),
        status_date=_parse_date(
            payload.get("CurrStatusDate"),
            result=result,
            row_number=row_number,
            field_name="CurrStatusDate",
        ),
        status_confidence=StatusConfidence.HIGH,
        status_source=source_name,
        date_delivery=_parse_date(
            payload.get("DeliveryDate"),
            result=result,
            row_number=row_number,
            field_name="DeliveryDate",
        ),
        planner_1_name=_clean_text(payload.get("Plan1Name")),
        planner_1_city=_clean_text(payload.get("Plan1City")),
        planner_1_email=_clean_text(payload.get("Plan1Email")),
        planner_1_phone=_clean_text(payload.get("Plan1Phone")),
        planner_2_name=_clean_text(payload.get("Plan2Name")),
        planner_2_city=_clean_text(payload.get("Plan2City")),
        planner_2_email=_clean_text(payload.get("Plan2Email")),
        planner_2_phone=_clean_text(payload.get("Plan2Phone")),
        source_urls=_dedupe_strings(_clean_text(payload.get(field)) for field in SOURCE_URL_FIELDS),
        last_editor=_clean_text(payload.get("Editor")),
        last_edit_date=_parse_date(
            payload.get("EditDate"),
            result=result,
            row_number=row_number,
            field_name="EditDate",
        ),
        created_by=PIPEDREAM_CREATED_BY,
    )

    project_notes = _build_project_notes(project, payload, imported_at)
    identifiers = _build_identifiers(project, payload, source_name, imported_at, project_identifier)
    status_history = _build_status_history(
        project,
        payload,
        source_name,
        result=result,
        row_number=row_number,
    )
    mapped_fields = {
        "project_name": project.project_name,
        "canonical_address": project.canonical_address,
        "pipeline_status": project.pipeline_status.value,
        "status_date": _serialize_json_value(project.status_date),
        "city": project.city,
        "state": project.state,
        "zip": project.zip,
        "total_units": project.total_units,
        "market_rate_units": project.market_rate_units,
        "affordable_units": project.affordable_units,
        "workforce_units": project.workforce_units,
        "developer": project.developer,
    }
    source_record = ProjectSourceRecord(
        project=project,
        source_name=source_name,
        source_record_id=project_identifier,
        source_url=project.source_urls[0] if project.source_urls else None,
        first_seen_at=imported_at,
        last_seen_at=imported_at,
        last_pulled_at=imported_at,
        raw_payload={
            key: _serialize_json_value(payload.get(key)) for key in HEADER_COLUMNS if key in payload
        },
        mapped_fields=mapped_fields,
        field_provenance={
            key: {"source": source_name, "confidence": StatusConfidence.HIGH.value}
            for key in mapped_fields
        },
    )

    return PipedreamProjectRecord(
        row_number=row_number,
        project_identifier_value=project_identifier,
        project=project,
        project_notes=project_notes,
        identifiers=identifiers,
        status_history=status_history,
        source_record=source_record,
    )


def _build_project_notes(
    project: Project,
    payload: dict[str, Any],
    imported_at: datetime,
) -> list[ProjectNote]:
    notes: list[ProjectNote] = []
    for source_field, note_type in {
        "Notes": "researcher_notes",
        "PersonalNotes": "personal_notes",
        "ChangeNotes": "change_notes",
    }.items():
        body = _clean_text(payload.get(source_field))
        if body is None:
            continue
        notes.append(
            ProjectNote(
                project=project,
                note_type=note_type,
                body=body,
                created_by_label=PIPEDREAM_CREATED_BY,
                created_at=imported_at,
            )
        )
    return notes


def _build_identifiers(
    project: Project,
    payload: dict[str, Any],
    source_name: str,
    imported_at: datetime,
    project_identifier: str,
) -> list[ProjectIdentifier]:
    identifiers = [
        ProjectIdentifier(
            project=project,
            identifier_type=IdentifierType.TCG_PIPEDREAM_ID,
            value=project_identifier,
            source=source_name,
            is_primary=True,
            first_seen_at=imported_at,
            last_seen_at=imported_at,
        )
    ]

    case_number = _clean_text(payload.get("RefNum"))
    if case_number:
        identifiers.append(
            ProjectIdentifier(
                project=project,
                identifier_type=IdentifierType.CASE_NUMBER,
                value=case_number,
                source=source_name,
                first_seen_at=imported_at,
                last_seen_at=imported_at,
            )
        )

    apn = _clean_identifier_text(payload.get("APN"))
    if apn:
        identifiers.append(
            ProjectIdentifier(
                project=project,
                identifier_type=IdentifierType.APN,
                value=apn,
                source=source_name,
                first_seen_at=imported_at,
                last_seen_at=imported_at,
            )
        )

    # Keep URL harvesting market-scoped until the seed ingester has a per-market extractor registry.
    if project.market == "los_angeles":
        for permit_number in extract_ladbs_pcis_permit_numbers(project.source_urls):
            identifiers.append(
                ProjectIdentifier(
                    project=project,
                    identifier_type=IdentifierType.PERMIT_NUMBER,
                    value=permit_number,
                    source=source_name,
                    first_seen_at=imported_at,
                    last_seen_at=imported_at,
                )
            )

    return identifiers


def _build_status_history(
    project: Project,
    payload: dict[str, Any],
    source_name: str,
    *,
    result: PipedreamImportResult,
    row_number: int,
) -> list[StatusHistory]:
    status_history: list[StatusHistory] = []
    for status_field, date_field in STATUS_FIELD_PAIRS:
        raw_status = _clean_text(payload.get(status_field))
        if not raw_status:
            continue
        status_history.append(
            StatusHistory(
                project=project,
                status=_parse_pipeline_status(
                    raw_status,
                    result=result,
                    row_number=row_number,
                    field_name=status_field,
                ),
                status_date=_parse_date(
                    payload.get(date_field),
                    result=result,
                    row_number=row_number,
                    field_name=date_field,
                ),
                source=source_name,
                notes=f"Imported from {status_field}",
            )
        )

    current_status = _clean_text(payload.get("CurrStatus"))
    if current_status:
        status_history.append(
            StatusHistory(
                project=project,
                status=_parse_pipeline_status(
                    current_status,
                    result=result,
                    row_number=row_number,
                    field_name="CurrStatus",
                ),
                status_date=_parse_date(
                    payload.get("CurrStatusDate"),
                    result=result,
                    row_number=row_number,
                    field_name="CurrStatusDate",
                ),
                source=source_name,
                notes="Imported from current Pipedream status",
            )
        )

    return status_history


def _build_relationships(
    payload: dict[str, Any],
    project_identifier: str,
    *,
    result: PipedreamImportResult,
    row_number: int,
) -> list[StagedProjectRelationship]:
    relationships: list[StagedProjectRelationship] = []
    for field_name, relationship_type in RELATIONSHIP_FIELD_MAP.items():
        related_identifier = _clean_project_identifier(
            payload.get(field_name),
            result=result,
            row_number=row_number,
            field_name=field_name,
        )
        if not related_identifier or related_identifier == project_identifier:
            continue
        relationships.append(
            StagedProjectRelationship(
                project_identifier_value=project_identifier,
                related_project_identifier_value=related_identifier,
                relationship_type=relationship_type,
                source_field=field_name,
                notes=f"Imported from {field_name}",
            )
        )
    return relationships


def _build_dismissed_record(
    *,
    payload: dict[str, Any],
    project_identifier: str,
    reason: DismissReason,
    source_name: str,
    market: str,
) -> DismissedRecord:
    address_value = _select_address_value(payload)
    normalized_address = normalize_address(
        address_value or "",
        city=_clean_text(payload.get("City")),
        state=_clean_text(payload.get("State")),
        postal_code=_clean_text(payload.get("Zip")),
        market=market,
    )
    corrp = _clean_project_identifier(payload.get("CorrP"))
    notes_parts = [_clean_text(payload.get("CurrStatus"))]
    if corrp:
        notes_parts.append(f"CorrP={corrp}")
    return DismissedRecord(
        source=source_name,
        source_record_id=project_identifier,
        canonical_address=normalized_address.canonical_address,
        reason=reason,
        dismissed_by=PIPEDREAM_CREATED_BY,
        notes="; ".join(part for part in notes_parts if part),
    )


def _select_address_value(payload: dict[str, Any]) -> str | None:
    address = _clean_text(payload.get("Address"))
    name = _clean_text(payload.get("Name"))
    if address and name and _looks_like_entity_name(address) and _looks_like_street_address(name):
        return name
    return address or name


def _looks_like_entity_name(value: str) -> bool:
    normalized = f" {_clean_text(value) or ''}".upper()
    return any(token in normalized for token in ENTITY_NAME_TOKENS)


def _looks_like_street_address(value: str) -> bool:
    normalized = (_clean_text(value) or "").upper()
    if not normalized or not ADDRESS_LIKE_RE.match(normalized):
        return False
    tokens = normalized.replace(".", " ").split()
    return any(token in ADDRESS_HINT_TOKENS for token in tokens)


def _parse_pipeline_status(
    value: Any,
    *,
    result: PipedreamImportResult | None = None,
    row_number: int | None = None,
    field_name: str = "CurrStatus",
) -> PipelineStatus:
    cleaned = _clean_text(value)
    if not cleaned:
        return PipelineStatus.PROPOSED

    status = PIPELINE_STATUS_MAP.get(cleaned)
    if status is not None:
        return status

    _record_issue(
        result=result,
        issue_type="invalid_status",
        row_number=row_number,
        field_name=field_name,
        raw_value=value,
        message="Unrecognized pipeline status; defaulting to Proposed",
    )
    return PipelineStatus.PROPOSED


def _parse_rent_or_sale(
    value: Any,
    *,
    result: PipedreamImportResult | None = None,
    row_number: int | None = None,
    field_name: str = "RentFS",
) -> RentOrSale:
    cleaned = _clean_text(value)
    if not cleaned:
        return RentOrSale.UNKNOWN

    parsed = RENT_OR_SALE_MAP.get(cleaned)
    if parsed is not None:
        return parsed

    _record_issue(
        result=result,
        issue_type="invalid_enum",
        row_number=row_number,
        field_name=field_name,
        raw_value=value,
        message="Unrecognized rent/sale value; defaulting to Unknown",
    )
    return RentOrSale.UNKNOWN


def _parse_product_type(
    value: Any,
    *,
    result: PipedreamImportResult | None = None,
    row_number: int | None = None,
    field_name: str = "ProdType",
) -> ProductType:
    cleaned = _clean_text(value)
    if not cleaned:
        return ProductType.UNKNOWN

    parsed = PRODUCT_TYPE_MAP.get(cleaned)
    if parsed is not None:
        return parsed

    _record_issue(
        result=result,
        issue_type="invalid_enum",
        row_number=row_number,
        field_name=field_name,
        raw_value=value,
        message="Unrecognized product type; defaulting to Other",
    )
    return ProductType.OTHER


def _parse_age_restriction(
    value: Any,
    *,
    result: PipedreamImportResult | None = None,
    row_number: int | None = None,
    field_name: str = "Senior",
) -> AgeRestriction:
    cleaned = _clean_text(value)
    if not cleaned:
        return AgeRestriction.UNKNOWN

    parsed = AGE_RESTRICTION_MAP.get(cleaned)
    if parsed is not None:
        return parsed

    _record_issue(
        result=result,
        issue_type="invalid_enum",
        row_number=row_number,
        field_name=field_name,
        raw_value=value,
        message="Unrecognized age restriction; defaulting to Unknown",
    )
    return AgeRestriction.UNKNOWN


def _clean_project_identifier(
    value: Any,
    *,
    result: PipedreamImportResult | None = None,
    row_number: int | None = None,
    field_name: str = "ProjectID",
) -> str | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None

    normalized = cleaned
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        normalized = f"{float(value):.5f}"
    elif "." in cleaned:
        prefix, suffix = cleaned.split(".", 1)
        if prefix.isdigit() and suffix.isdigit():
            normalized = f"{prefix}.{suffix.ljust(5, '0')[:5]}"

    if _is_suspicious_project_identifier(value, cleaned):
        _record_issue(
            result=result,
            issue_type="suspicious_identifier",
            row_number=row_number,
            field_name=field_name,
            raw_value=value,
            message=(
                f"Project identifier normalized to {normalized!r}; verify abbreviated ID format"
            ),
        )

    return normalized


def _parse_date(
    value: Any,
    *,
    result: PipedreamImportResult | None = None,
    row_number: int | None = None,
    field_name: str = "date",
) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    cleaned = _clean_text(value)
    if cleaned is None:
        return None

    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue

    _record_issue(
        result=result,
        issue_type="invalid_date",
        row_number=row_number,
        field_name=field_name,
        raw_value=value,
        message="Unrecognized date value; storing null",
    )
    return None


def _record_issue(
    *,
    result: PipedreamImportResult | None,
    issue_type: str,
    row_number: int | None,
    field_name: str,
    raw_value: Any,
    message: str,
) -> None:
    if result is None or row_number is None:
        return
    result.add_issue(
        issue_type=issue_type,
        row_number=row_number,
        field_name=field_name,
        raw_value=raw_value,
        message=message,
    )


def _is_suspicious_project_identifier(value: Any, cleaned: str) -> bool:
    if "." not in cleaned:
        return False

    prefix, suffix = cleaned.split(".", 1)
    if not prefix.isdigit() or not suffix.isdigit():
        return False
    if len(suffix) >= 5:
        return False

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    return cleaned != f"{prefix}.{suffix.ljust(5, '0')[:5]}"
