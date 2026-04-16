from __future__ import annotations

import logging
from collections import Counter
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from geoalchemy2.elements import WKTElement
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from tcg_pipeline.db.models import (
    AgeRestriction,
    GeocodeConfidence,
    IdentifierType,
    PipelineStatus,
    ProductType,
    Project,
    ProjectIdentifier,
    ProjectSourceRecord,
    RentOrSale,
    StatusConfidence,
    StatusHistory,
)
from tcg_pipeline.matching.normalizer import (
    normalize_address,
    normalize_city,
    normalize_postal_code,
)

logger = logging.getLogger(__name__)

COSTAR_SOURCE_NAME = "costar"
COSTAR_CREATED_BY = "costar_import"
HEADER_ROW_INDEX = 1
DATA_START_ROW_INDEX = 2
NULL_SENTINELS = {"", "--"}
MONTH_NAME_TO_NUMBER = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "SEPT": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}
CONSTR_STATUS_MAP = {
    "UNDER CONSTRUCTION": PipelineStatus.UNDER_CONSTRUCTION,
    "FINAL PLANNING": PipelineStatus.APPROVED,
    "PROPOSED": PipelineStatus.PROPOSED,
    "DEFERRED": PipelineStatus.STALLED,
    "ABANDONED": PipelineStatus.INACTIVE,
}


@dataclass(slots=True)
class CoStarImportIssue:
    issue_type: str
    workbook_name: str
    row_number: int
    field_name: str
    raw_value: str | None
    message: str


@dataclass(slots=True)
class CoStarProjectRecord:
    workbook_name: str
    row_number: int
    property_id: str
    project: Project
    identifiers: list[ProjectIdentifier]
    status_history: list[StatusHistory]
    source_record: ProjectSourceRecord


@dataclass(slots=True)
class CoStarImportResult:
    source_paths: list[Path]
    project_records: list[CoStarProjectRecord] = field(default_factory=list)
    skipped_property_ids: list[str] = field(default_factory=list)
    duplicate_property_ids: list[str] = field(default_factory=list)
    issues: list[CoStarImportIssue] = field(default_factory=list)
    missing_property_id_rows: int = 0

    @property
    def imported_count(self) -> int:
        return len(self.project_records)

    @property
    def duplicate_count(self) -> int:
        return len(self.duplicate_property_ids)

    @property
    def issue_count(self) -> int:
        return len(self.issues)

    @property
    def issue_counts(self) -> dict[str, int]:
        return dict(Counter(issue.issue_type for issue in self.issues))

    def add_issue(
        self,
        *,
        issue_type: str,
        workbook_name: str,
        row_number: int,
        field_name: str,
        raw_value: Any,
        message: str,
    ) -> None:
        issue = CoStarImportIssue(
            issue_type=issue_type,
            workbook_name=workbook_name,
            row_number=row_number,
            field_name=field_name,
            raw_value=_display_value(raw_value),
            message=message,
        )
        if issue in self.issues:
            return
        self.issues.append(issue)
        logger.warning(
            "%s row %s field %s: %s (raw=%r)",
            workbook_name,
            row_number,
            field_name,
            message,
            issue.raw_value,
        )


class CoStarIngester:
    def __init__(
        self,
        *,
        market: str,
        source_name: str = COSTAR_SOURCE_NAME,
        allowed_cities: Iterable[str] | None = None,
    ) -> None:
        self.market = market
        self.source_name = source_name
        self.allowed_cities = {
            normalized
            for city in (allowed_cities or [])
            if (normalized := normalize_city(city, market=market))
        }

    def ingest_workbook(self, workbook_path: str | Path) -> CoStarImportResult:
        return self.ingest_workbooks([workbook_path])

    def ingest_workbooks(self, workbook_paths: Sequence[str | Path]) -> CoStarImportResult:
        result = CoStarImportResult(source_paths=[Path(path) for path in workbook_paths])
        seen_property_ids: set[str] = set()

        for source_path in result.source_paths:
            workbook = load_workbook(source_path, data_only=True, read_only=True)
            try:
                worksheet = workbook[workbook.sheetnames[0]]
                header_map = _extract_headers(worksheet)
                imported_at = datetime.now(UTC)

                for row_number, row in enumerate(
                    worksheet.iter_rows(min_row=DATA_START_ROW_INDEX),
                    start=DATA_START_ROW_INDEX,
                ):
                    payload = _build_row_payload(row, header_map)
                    property_id = _clean_text(payload.get("PropertyID"))
                    if not property_id:
                        if _row_has_values(payload):
                            result.missing_property_id_rows += 1
                            logger.warning(
                                "%s row %s skipped: missing PropertyID",
                                source_path.name,
                                row_number,
                            )
                        continue

                    if property_id in seen_property_ids:
                        result.duplicate_property_ids.append(property_id)
                        continue

                    normalized_city = normalize_city(payload.get("City"), market=self.market)
                    if self.allowed_cities and normalized_city not in self.allowed_cities:
                        result.skipped_property_ids.append(property_id)
                        continue

                    seen_property_ids.add(property_id)
                    project_record = _build_project_record(
                        payload=payload,
                        workbook_name=source_path.name,
                        row_number=row_number,
                        property_id=property_id,
                        market=self.market,
                        source_name=self.source_name,
                        imported_at=imported_at,
                        result=result,
                    )
                    result.project_records.append(project_record)
            finally:
                workbook.close()

        return result


def _extract_headers(worksheet: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in worksheet[HEADER_ROW_INDEX]:
        value = _clean_text(cell.value)
        if value:
            headers[value] = cell.column
    return headers


def _build_row_payload(row: tuple[Any, ...], header_map: dict[str, int]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for header, column_index in header_map.items():
        if column_index > len(row):
            continue
        payload[header] = row[column_index - 1].value
    return payload


def _build_project_record(
    *,
    payload: dict[str, Any],
    workbook_name: str,
    row_number: int,
    property_id: str,
    market: str,
    source_name: str,
    imported_at: datetime,
    result: CoStarImportResult,
) -> CoStarProjectRecord:
    normalized_address = normalize_address(
        _clean_text(payload.get("Property Address")) or "",
        city=_clean_text(payload.get("City")),
        state=_clean_text(payload.get("State")),
        postal_code=_clean_text(payload.get("Zip")),
        market=market,
    )
    property_type_raw = _clean_text(payload.get("Property Type"))
    secondary_type = _clean_text(payload.get("Secondary Type"))

    project = Project(
        canonical_address=normalized_address.canonical_address or normalized_address.raw_address,
        raw_addresses=_dedupe_strings([_clean_text(payload.get("Property Address"))]),
        lat=_parse_float(payload.get("Latitude")),
        lng=_parse_float(payload.get("Longitude")),
        location=_build_location(payload.get("Latitude"), payload.get("Longitude")),
        geocode_confidence=_determine_geocode_confidence(payload),
        market=market,
        city=normalized_address.city or normalize_city(payload.get("City"), market=market) or "",
        state=normalized_address.state or _clean_text(payload.get("State")) or "",
        county=_clean_text(payload.get("County Name")) or "",
        zip=normalize_postal_code(payload.get("Zip")),
        costar_submarket=_clean_text(payload.get("Submarket Name")),
        zoning=_clean_text(payload.get("Zoning")),
        project_name=_clean_text(payload.get("Property Name")),
        developer=_clean_text(payload.get("Developer Name")),
        rent_or_sale=_parse_rent_or_sale(
            property_type_raw=property_type_raw,
            secondary_type=secondary_type,
            rent_type=_clean_text(payload.get("Rent Type")),
        ),
        product_type=_parse_product_type(
            property_type_raw=property_type_raw,
            secondary_type=secondary_type,
        ),
        age_restriction=_parse_age_restriction(
            payload.get("Market Segment"),
            result=result,
            workbook_name=workbook_name,
            row_number=row_number,
        ),
        stories=_parse_int(payload.get("Number Of Stories")),
        total_units=_parse_int(payload.get("Number Of Units")),
        pct_studio=_parse_bed_mix(
            payload.get("% Studios"),
            result=result,
            workbook_name=workbook_name,
            row_number=row_number,
            field_name="% Studios",
        ),
        pct_1bed=_parse_bed_mix(
            payload.get("% 1-Bed"),
            result=result,
            workbook_name=workbook_name,
            row_number=row_number,
            field_name="% 1-Bed",
        ),
        pct_2bed=_parse_bed_mix(
            payload.get("% 2-Bed"),
            result=result,
            workbook_name=workbook_name,
            row_number=row_number,
            field_name="% 2-Bed",
        ),
        pct_other_bed=_combine_other_bed_mix(
            payload.get("% 3-Bed"),
            payload.get("% 4-Bed"),
            result=result,
            workbook_name=workbook_name,
            row_number=row_number,
        ),
        acres=_parse_float(payload.get("Land Area (AC)")),
        hotel_keys=_parse_int(payload.get("Rooms")),
        total_sf=_parse_int(payload.get("RBA")),
        parking_spaces=_parse_int(payload.get("Number Of Parking Spaces")),
        style=_clean_text(payload.get("Style")),
        property_type=property_type_raw,
        affordable_type=_clean_text(payload.get("Affordable Type")),
        owner=_clean_text(payload.get("Owner Name")),
        true_owner=_clean_text(payload.get("True Owner Name")),
        architect=_clean_text(payload.get("Architect Name")),
        pipeline_status=_parse_pipeline_status(
            payload.get("Constr Status"),
            result=result,
            workbook_name=workbook_name,
            row_number=row_number,
        ),
        status_confidence=StatusConfidence.MEDIUM,
        status_source=source_name,
        date_delivery=_parse_delivery_date(
            payload.get("Year Built"),
            payload.get("Month Built"),
            result=result,
            workbook_name=workbook_name,
            row_number=row_number,
        ),
        date_construction_start=_parse_month_year_date(
            payload.get("Construction Begin"),
            result=result,
            workbook_name=workbook_name,
            row_number=row_number,
            field_name="Construction Begin",
        ),
        created_by=COSTAR_CREATED_BY,
    )

    identifiers = _build_identifiers(project, payload, source_name, imported_at, property_id)
    status_history = [
        StatusHistory(
            project=project,
            status=project.pipeline_status,
            status_date=project.date_construction_start or project.date_delivery,
            source=source_name,
            notes="Imported from CoStar Constr Status",
        )
    ]
    mapped_fields = {
        "project_name": project.project_name,
        "canonical_address": project.canonical_address,
        "city": project.city,
        "state": project.state,
        "zip": project.zip,
        "pipeline_status": project.pipeline_status.value,
        "date_construction_start": _serialize_json_value(project.date_construction_start),
        "date_delivery": _serialize_json_value(project.date_delivery),
        "property_type": project.property_type,
        "total_units": project.total_units,
        "total_sf": project.total_sf,
        "developer": project.developer,
        "owner": project.owner,
    }
    source_record = ProjectSourceRecord(
        project=project,
        source_name=source_name,
        source_record_id=property_id,
        first_seen_at=imported_at,
        last_seen_at=imported_at,
        last_pulled_at=imported_at,
        raw_payload={key: _serialize_json_value(value) for key, value in payload.items()},
        mapped_fields=mapped_fields,
        field_provenance={
            key: {"source": source_name, "confidence": StatusConfidence.MEDIUM.value}
            for key in mapped_fields
        },
    )

    return CoStarProjectRecord(
        workbook_name=workbook_name,
        row_number=row_number,
        property_id=property_id,
        project=project,
        identifiers=identifiers,
        status_history=status_history,
        source_record=source_record,
    )


def _build_identifiers(
    project: Project,
    payload: dict[str, Any],
    source_name: str,
    imported_at: datetime,
    property_id: str,
) -> list[ProjectIdentifier]:
    identifiers = [
        ProjectIdentifier(
            project=project,
            identifier_type=IdentifierType.COSTAR_PROPERTY_ID,
            value=property_id,
            source=source_name,
            is_primary=True,
            first_seen_at=imported_at,
            last_seen_at=imported_at,
        )
    ]
    seen_identifier_keys = {
        (identifiers[0].identifier_type, identifiers[0].value),
    }
    for field_name in ("Parcel Number 1(Min)", "Parcel Number 2(Max)"):
        apn = _clean_identifier_text(payload.get(field_name))
        if not apn:
            continue
        identifier_key = (IdentifierType.APN, apn)
        if identifier_key in seen_identifier_keys:
            continue
        seen_identifier_keys.add(identifier_key)
        identifiers.append(
            ProjectIdentifier(
                project=project,
                identifier_type=IdentifierType.APN,
                value=apn,
                source=source_name,
                first_seen_at=imported_at,
                last_seen_at=imported_at,
            )
        )
    return identifiers


def _parse_pipeline_status(
    value: Any,
    *,
    result: CoStarImportResult,
    workbook_name: str,
    row_number: int,
) -> PipelineStatus:
    cleaned = _clean_text(value)
    if not cleaned:
        return PipelineStatus.PROPOSED

    status = CONSTR_STATUS_MAP.get(cleaned.upper())
    if status is not None:
        return status

    result.add_issue(
        issue_type="invalid_status",
        workbook_name=workbook_name,
        row_number=row_number,
        field_name="Constr Status",
        raw_value=value,
        message="Unrecognized CoStar construction status; defaulting to Proposed",
    )
    return PipelineStatus.PROPOSED


def _parse_rent_or_sale(
    *,
    property_type_raw: str | None,
    secondary_type: str | None,
    rent_type: str | None,
) -> RentOrSale:
    property_type = (property_type_raw or "").upper()
    secondary = (secondary_type or "").upper()
    if "CONDO" in property_type or "CONDO" in secondary:
        return RentOrSale.FOR_SALE
    if rent_type:
        return RentOrSale.RENTAL
    return RentOrSale.UNKNOWN


def _parse_product_type(
    *,
    property_type_raw: str | None,
    secondary_type: str | None,
) -> ProductType:
    property_type = (property_type_raw or "").upper()
    secondary = (secondary_type or "").upper()
    if property_type == "MULTIFAMILY":
        if "CONDO" in secondary:
            return ProductType.CONDO
        if "TOWNHOME" in secondary:
            return ProductType.TOWNHOME
        if "MICRO" in secondary or "CO-LIVING" in secondary:
            return ProductType.MICRO_CO_LIVING
        return ProductType.APARTMENT
    return ProductType.UNKNOWN


def _parse_age_restriction(
    value: Any,
    *,
    result: CoStarImportResult,
    workbook_name: str,
    row_number: int,
) -> AgeRestriction:
    cleaned = (_clean_text(value) or "").upper()
    if not cleaned:
        return AgeRestriction.UNKNOWN
    if cleaned == "SENIOR":
        return AgeRestriction.SENIOR
    if cleaned == "STUDENT":
        return AgeRestriction.STUDENT
    if cleaned in {"ALL", "CORPORATE", "MILITARY"}:
        return AgeRestriction.NON_AGE_RESTRICTED

    result.add_issue(
        issue_type="invalid_enum",
        workbook_name=workbook_name,
        row_number=row_number,
        field_name="Market Segment",
        raw_value=value,
        message="Unrecognized market segment; defaulting age restriction to Unknown",
    )
    return AgeRestriction.UNKNOWN


def _parse_bed_mix(
    value: Any,
    *,
    result: CoStarImportResult,
    workbook_name: str,
    row_number: int,
    field_name: str,
) -> float | None:
    numeric = _parse_float(value)
    if numeric is None:
        return None
    if numeric < 0 or numeric > 100:
        result.add_issue(
            issue_type="invalid_percentage",
            workbook_name=workbook_name,
            row_number=row_number,
            field_name=field_name,
            raw_value=value,
            message="Bed mix percentage outside 0-100 range; storing normalized float anyway",
        )
    return round(numeric / 100.0, 6)


def _combine_other_bed_mix(
    three_bed: Any,
    four_bed: Any,
    *,
    result: CoStarImportResult,
    workbook_name: str,
    row_number: int,
) -> float | None:
    values = [
        _parse_bed_mix(
            three_bed,
            result=result,
            workbook_name=workbook_name,
            row_number=row_number,
            field_name="% 3-Bed",
        ),
        _parse_bed_mix(
            four_bed,
            result=result,
            workbook_name=workbook_name,
            row_number=row_number,
            field_name="% 4-Bed",
        ),
    ]
    non_null_values = [value for value in values if value is not None]
    if not non_null_values:
        return None
    return round(sum(non_null_values), 6)


def _parse_month_year_date(
    value: Any,
    *,
    result: CoStarImportResult,
    workbook_name: str,
    row_number: int,
    field_name: str,
) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)

    cleaned = _clean_text(value)
    if not cleaned:
        return None
    parts = cleaned.replace(",", " ").split()
    if len(parts) != 2:
        result.add_issue(
            issue_type="invalid_date",
            workbook_name=workbook_name,
            row_number=row_number,
            field_name=field_name,
            raw_value=value,
            message="Unrecognized month/year date; storing null",
        )
        return None
    month = _parse_month_token(parts[0])
    year = _parse_int(parts[1])
    if month is None or year is None:
        result.add_issue(
            issue_type="invalid_date",
            workbook_name=workbook_name,
            row_number=row_number,
            field_name=field_name,
            raw_value=value,
            message="Unrecognized month/year date; storing null",
        )
        return None
    return date(year, month, 1)


def _parse_delivery_date(
    year_value: Any,
    month_value: Any,
    *,
    result: CoStarImportResult,
    workbook_name: str,
    row_number: int,
) -> date | None:
    year = _parse_int(year_value)
    if year is None:
        return None
    month = _parse_month_component(month_value)
    if month is None:
        month = 1
    if month < 1 or month > 12:
        result.add_issue(
            issue_type="invalid_date",
            workbook_name=workbook_name,
            row_number=row_number,
            field_name="Month Built",
            raw_value=month_value,
            message="Invalid month built value; defaulting date_delivery month to January",
        )
        month = 1
    return date(year, month, 1)


def _parse_month_component(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value)
    cleaned = _clean_text(value)
    if not cleaned:
        return None
    if cleaned.isdigit():
        return int(cleaned)
    return _parse_month_token(cleaned)


def _parse_month_token(value: str) -> int | None:
    alpha_only = "".join(character for character in value.upper() if character.isalpha())
    if not alpha_only:
        return None
    if alpha_only in MONTH_NAME_TO_NUMBER:
        return MONTH_NAME_TO_NUMBER[alpha_only]
    return MONTH_NAME_TO_NUMBER.get(alpha_only[:3])


def _build_location(lat_value: Any, lng_value: Any) -> WKTElement | None:
    lat = _parse_float(lat_value)
    lng = _parse_float(lng_value)
    if lat is None or lng is None:
        return None
    return WKTElement(f"POINT({lng} {lat})", srid=4326)


def _determine_geocode_confidence(payload: dict[str, Any]) -> GeocodeConfidence:
    latitude = _parse_float(payload.get("Latitude"))
    longitude = _parse_float(payload.get("Longitude"))
    if latitude is not None and longitude is not None:
        return GeocodeConfidence.HIGH
    return GeocodeConfidence.NONE


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if text in NULL_SENTINELS:
        return None
    return text


def _clean_identifier_text(value: Any) -> str | None:
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = _parse_float(value)
        if numeric is None:
            return cleaned
        return str(int(numeric)) if numeric.is_integer() else str(numeric)
    return cleaned


def _parse_int(value: Any) -> int | None:
    numeric = _parse_float(value)
    if numeric is None:
        return None
    return int(round(numeric))


def _parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    cleaned = _clean_text(value)
    if cleaned is None:
        return None

    normalized = cleaned.replace(",", "").replace("%", "")
    try:
        return float(normalized)
    except ValueError:
        return None


def _dedupe_strings(values: Iterable[str | None]) -> list[str]:
    deduped: list[str] = []
    for value in values:
        if not value or value in deduped:
            continue
        deduped.append(value)
    return deduped


def _display_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _serialize_json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _row_has_values(payload: dict[str, Any]) -> bool:
    return any(
        value is not None and str(value).strip() not in NULL_SENTINELS
        for value in payload.values()
    )
